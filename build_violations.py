from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# ── Color palette ─────────────────────────────────────────────────────────────
HDR_DARK   = "1F3864"
HDR_MED    = "2E75B6"
RED_BG     = "FCE4D6"
RED_TXT    = "C00000"
GOLD_BG    = "FFF2CC"
GOLD_TXT   = "7F6000"
GREEN_BG   = "E2EFDA"
GREEN_TXT  = "375623"
PURPLE_BG  = "EAD1DC"
PURPLE_TXT = "6B1F42"
BLUE_BG    = "DDEEFF"
BLUE_TXT   = "1F3864"
GRAY_BG    = "F2F2F2"
WHITE      = "FFFFFF"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(color="FFFFFF", sz=11, bold=True):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def cell_font(color="000000", sz=10, bold=False, italic=False):
    return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def wrap():
    return Alignment(wrap_text=True, vertical="top")

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — VIOLATIONS DATABASE
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Violations Database"
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 36

headers = [
    "ID", "Date", "Category", "Type",
    "Severity", "Person(s) Involved",
    "Rule / Statute / Standard Violated",
    "Description of Violation",
    "Key Evidence",
    "Legal Exposure",
    "Status"
]

# Header row
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = hdr_font()
    c.fill = fill(HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# ── Data rows ────────────────────────────────────────────────────────────────
# Format: (id, date, category, type, severity, people, rule, description, evidence, exposure, status)
# Severity: CRITICAL / HIGH / MEDIUM / LOW
# Type: Legal / Ethics / Governance / Procedural / Financial / Conflict of Interest
rows = [

    # ── FINANCIAL / FIDUCIARY ──────────────────────────────────────────────
    ("F-01", "Mar 17, 2025; Oct 29–31, 2025", "Financial Records Refusal", "Legal + Governance",
     "CRITICAL", "Jeanne Grealish",
     "NMSA §53-8-26 / §53-8-27 (Nonprofit Records Inspection); PMTNM Handbook EST duties: 'The records should be open to any member at all times'",
     "TWO SEPARATE REFUSALS documented:\n\n(1) Mar 17, 2025: VP formally requests financial records in his capacity as VP and Budget Committee member. Grealish refuses, claiming VP 'has no function beyond August.' Sharon Kunitz reinforces the refusal. This is the core legal violation that precipitated the petition.\n\n(2) Oct 29–31, 2025: Attorney Kathy Black (Law 4 Small Business) sends formal demand letter to Grealish under NMSA §53-8-27, requesting all financial records for FY ended June 30, 2025, by close of business October 31. Grealish's private response to Kathy: 'It is not physically possible for me to bring the requested items to your office on Friday, October 31, 2025.' Grealish also confirms: 'PMTNM does not do online banking' — no electronic production is possible either. Records refused to VP's attorney within the statutory deadline.",
     "Email Mar 17, 2025 (Grealish → Miller); Sharon Kunitz CC reply; Kathy Black demand letter Oct 29, 2025 citing NMSA §53-8-27 (Gmail thread ID: 19a3309ed58127ac); Grealish private letter to Kathy Black Oct 30, 2025: 'It is not physically possible for me to bring the requested items to your office on Friday, October 31, 2025' and 'PMTNM does not do online banking'; Grealish's own prior statement: 'All checkbooks...available for inspection by any member upon request'",
     "Two documented refusals — first to VP directly, second to VP's attorney under formal statutory demand. Both refused. Attorney's demand cited NMSA §53-8-27 explicitly. Grealish's 'not physically possible' response to a statutory demand is itself evidence of the records-control structure that the petition seeks to remedy.",
     "Active — Primary basis of litigation; attorney demand also refused"),

    ("F-02", "Oct 7, 2025", "Budget Committee Obstruction", "Legal + Governance",
     "CRITICAL", "Jeanne Grealish",
     "PMTNM Handbook (2004, 2014): Bylaws Art. IX §2 — VP 'shall be a member of the Budget Committee'; Section VI-B — 'Budget Committee presents a proposed budget for the new fiscal year for approval by the Executive Board at the Conference Board Meeting'; EST duties — 'works with the Budget Committee to formulate the fiscal year budget'",
     "Chip formally requested a budget committee meeting multiple times throughout 2025. On Oct 7, 2025, Grealish responded in writing: 'There probably will be no need for an actual meeting.' In the same email, Grealish directly dismissed the handbook provisions Chip was invoking: 'As for any statements in the Handbook regarding the budget, committee, dates, etc. these would be grossly out of date since the Handbook itself is no longer current and requires serious revision. Thanks heaven, PMTNM does not function under any legal mandates!' This is the DIRECT CONTEXT of the 'no legal mandates' statement — it was made specifically to shut down Chip's invocation of the Budget Committee requirement.\n\nPMTNM's fiscal year began July 1, 2025. The CPA report was not received until Oct 26. The proposed budget was not sent to the committee until Oct 29. Instead of a proper meeting, Grealish substituted email exchange. The budget was then 'voted on' at the Nov 7 board meeting — more than 4 months into the fiscal year, with no approved budget in place during that period. During those 4+ months, Grealish continued spending organizational funds (storage, honorarium, operational expenses) under no approved budget.\n\nKathy Black (Chip's attorney) reviewed the situation and concluded: 'the organization is not following the handbook regarding committees.' She also noted at this time that 'the State of New Mexico annual report has not been filed.'",
     "Oct 7, 2025 Grealish email to Chip + Sharon + Laura + Larry + Heather Nasi (thread ID: 199c0d60237afa6c): 'There probably will be no need for an actual meeting'; 'the Handbook itself is no longer current and requires serious revision. Thanks heaven, PMTNM does not function under any legal mandates!'; Oct 29, 2025 Grealish email distributing budget (thread ID: 19a3309ed58127ac): budget sent 4 months into fiscal year; Oct 30, 2025 Kathy Black email to Chip: 'the organization is not following the handbook regarding committees'; 2004 Handbook Bylaws Art. IX §2 (VP on Budget Committee); 2004 Handbook Section VI-B (budget requires formal Executive Board approval at board meeting)",
     "No budget committee meeting held. Handbook requirements for budget approval dismissed with 'no legal mandates' statement. Budget sent via email 4+ months into fiscal year. Grealish spent organizational funds throughout July–November with no approved budget. Kathy Black confirmed handbook violation. NM annual report also not filed as of late October 2025.",
     "Email-verified + Attorney-confirmed — 'No legal mandates' CONTEXT established"),

    ("F-03", "Oct 25, 2025", "Financial Control", "Governance",
     "HIGH", "Jeanne Grealish",
     "Standard nonprofit governance; PMTNM Handbook on financial reporting timelines",
     "CPA FY 2024-2025 report distributed 5 months after fiscal year end. Budget for 2025-2026 proposed after the fiscal year had already begun (July 1, 2025). Grealish confused about which year the budget applies to (confirmed on Nov 4 transcript).",
     "Email Oct 25 (Grealish → Board with CPA report); Nov 4 transcript (Kathy Black: 'I don't understand why you waited five months'); Budget documents (DocsFromJeanne)",
     "Operating without an approved budget for 5+ months of fiscal year. Kathy Black: potential State of NM compliance issue.",
     "Documented"),

    ("F-04", "Nov 7, 2025", "Financial Control", "Legal + Procedural",
     "CRITICAL", "Jeanne Grealish; Sharon Kunitz (as presiding officer)",
     "Robert's Rules §49 (voting requirements); Standard nonprofit governance — financial reports require motion + vote to approve",
     "Board minutes state financial report 'was approved as presented.' Audio recording of the Nov 7 meeting confirms NO motion was made, NO second was given, and NO vote was taken to approve the financial report. Only a separate motion about CPA-designated accounts was carried — minutes conflate the two.",
     "Nov 7 Board Meeting Audio Recording; Nov 7 Transcript; INTERNAL_Minutes Review (audio-timestamped); Proposed Corrections to Minutes (Apr 10, 2026)",
     "Potential falsification of board minutes. Any 'approval' is procedurally void. Pattern of treating financial matters as rubber-stamp items.",
     "Audio-verified — Key evidence for litigation"),

    ("F-05", "Ongoing 2024-2026", "Financial Control", "Legal + Ethics",
     "CRITICAL", "Jeanne Grealish",
     "NM Nonprofit Corporation Act; Fiduciary duty; Basic nonprofit governance standards",
     "Grealish is sole signatory on ALL PMTNM bank accounts, sole contact with CPA, sole recipient of MTNA membership updates, sole keeper of all financial records. No board member has independent financial access. Organization operates from her private residence. Outgoing treasurer Larry Blind explicitly warned this arrangement was improper (Dec 2024).",
     "Kathy Black Initial Impressions (Oct 2025); Larry Blind bank warning (Dec 2024); Nov 4 transcript; DocsFromJeanne financial records; Bank statements showing sole signatory",
     "Concentration of financial control in one unaccountable 20-year incumbent. Creates conditions for embezzlement, self-dealing, and organizational failure if she becomes incapacitated.",
     "Structural — Core to petition"),

    ("F-06", "Nov 4, 2025", "Financial Self-Dealing", "Ethics + Financial",
     "HIGH", "Jeanne Grealish",
     "IRS private benefit prohibition (501(c)(6)); Fiduciary duty; Nonprofit self-dealing rules",
     "Grealish admitted in front of attorney Kathy Black that she stores personal furniture in the PMTNM storage unit, which costs approximately $266/month. The proposed FY budget allocates ~30% to storage. Chip's document to Sam and Laura calculates 65% of the budget goes to storage and Grealish-related expenses.",
     "Nov 4, 2025 meeting transcript (Grealish admission before Kathy Black); DocsFromJeanne budget documents; Chip's budget analysis to Sam/Laura; Kathy Black present as witness",
     "Using nonprofit resources for personal benefit. IRS private benefit rules prohibit this for 501(c)(6) orgs. Kathy Black witnessed the admission — she is a potential witness.",
     "Attorney-witnessed — Strong evidence"),

    ("F-07", "Ongoing 2024-2026", "Financial Self-Dealing", "Ethics + Financial",
     "MEDIUM", "Jeanne Grealish",
     "Nonprofit executive compensation standards; PMTNM Handbook on EST honorarium",
     "Grealish receives an EST honorarium of $1,750/year with no documented performance measures, no board evaluation process, and no defined scope of work tied to payment. Honorarium continues despite documented administrative failures.",
     "DocsFromJeanne financial statements; FY 2024-2025 CPA report showing $1,750 EST honorarium",
     "Honorarium without performance standards creates appearance of self-dealing. Board has no mechanism to reduce or revoke honorarium.",
     "Documented"),

    ("F-08", "Apr 28, 2025", "Unsupported Negative Characterization", "Governance + Ethics",
     "MEDIUM", "Jeanne Grealish",
     "Board fiduciary duty; Professional conduct; Accurate characterization of member contributions",
     "During TOTY award discussion, Grealish made an unsupported negative characterization of Teri Reck's contributions: 'I believe the reason Terry Rett was not posted last time was she has not contributed very much to community service.' This claim was immediately contradicted by board members Chip Miller and Sharon Kunitz, who pointed to Reck's actual contributions: workshop leadership, conference activities, fundraising, and 47 years of teaching. The comment demonstrates problematic board governance — unsupported negative statements made during award decision processes can marginalize members and skew institutional recognition. See F-11 for pattern context (back-to-back TOTY suppression cycles).",
     "April 28, 2025 Board Meeting Audio & Transcript: timestamp 01:39:00–01:39:35 (Grealish's claim at 01:39:00; Chip's contradiction at 01:39:35); Transcript lines 1851–1881 (redundant due to audio-to-text conversion artifacts); Teri Reck's own 2023 TOTY submission (PDF: Terri Reck.pdf, Gmail thread ID: 19239a27e03b7010) documenting actual contributions: 47 years teaching, Dalcroze certification, SF/NM Symphony harpsichordist, Music Bowl Chair, PMTNM silent auction donations",
     "Unsupported negative characterization made during official board decision-making (TOTY award discussion). Immediately contradicted by fellow board members. Pattern evidence in F-11 shows back-to-back cycles of suppression targeting same individual.",
     "Transcript-verified — Substantive challenge to characterization"),

    ("F-09", "Apr 25, 2025", "Financial Irregularity", "Governance",
     "LOW", "Jeanne Grealish",
     "Basic nonprofit communications standards; Member value proposition",
     "Grealish stated she didn't want to work on the membership booklet, citing that it 'costs 90 cents to mail.' A trivial cost objection used to block a core member-service function while simultaneously budgeting 30%+ for personal storage.",
     "Spring 2025 Board Meeting transcript; Budget documents showing storage vs. communications allocations",
     "Demonstrates misaligned organizational priorities and selective cost objections.",
     "Documented — Context evidence"),

    # ── LARRY BLIND — CONFLICTS OF INTEREST ───────────────────────────────
    ("L-01", "Oct 4, 2023", "Conflict of Interest / Boss as Nominee", "Ethics + Procedural",
     "CRITICAL", "Larry Blind",
     "MTNA Code of Ethics §3 (conflicts of interest); Nonprofit governance — board member conflicts; Robert's Rules — impartiality; NM nonprofit law — fiduciary duty",
     "Larry Blind, an employee of NM School of Music owned by Tatiana Vetrinskaya, sent the Teacher of the Year voting solicitation from his NMSM business email (lblind@nmschoolofmusic.com) on behalf of PMTNM. The winning nominee was Tatiana Vetrinskaya (tatiana@nmschoolofmusic.com) — his employer and NMSM owner. Larry controls Tatiana's employment at NMSM while administering a process at PMTNM that determines whether she receives professional honors. He cannot be impartial about voting for or administering a process involving his boss. The email solicitation came from his business email address, blurring lines between his employer's resources and PMTNM business.",
     "Oct 4, 2023 email from lblind@nmschoolofmusic.com soliciting TOTY votes (forwarded to Chip by Haewon Yang Sep 28, 2024); Tatiana's email address tatiana@nmschoolofmusic.com; NMSM organizational structure showing Tatiana as co-founder/owner; Haewon thread 'FTC guidelines for MTNA and PMTNM' Sep 27-28, 2024; Board Minutes and NMSM documentation showing Larry as NMSM employee",
     "An employee controlling a voting process that determines whether his boss receives professional honors. Maximum conflict of interest: Larry's employment and livelihood depend on Tatiana's goodwill, making impartial administration impossible. Used business email to conduct PMTNM business, blurring organizational boundaries.",
     "Email-verified — Relationship confirmed in NMSM documentation"),

    ("L-02", "Nov 7-8, 2023", "Conflict of Interest", "Ethics + Legal",
     "HIGH", "Larry Blind",
     "MTNA Code of Ethics; Nonprofit conference standards — conference should serve members' students, not board members' commercial interests",
     "PMTNM conference featured students from NM School of Music (Larry's business) who were not necessarily studying with PMTNM members. PMTNM members paid to attend and were expected to sit through performances by students of a competing commercial entity whose owner serves on the PMTNM board.",
     "Nov 14, 2023 ethics complaint to MTNA (Gary Ingle); Conference program/documentation from 2023; Haewon thread referencing 'the conference last year wasn't the best'",
     "Using PMTNM's annual conference as a marketing vehicle for Larry's private music school. Dilutes conference value for members and creates unfair commercial advantage.",
     "First ethics complaint filed Nov 2023 — no action taken"),

    ("L-03", "Nov-Dec 2024", "Conflict of Interest", "Governance",
     "HIGH", "Larry Blind",
     "Nonprofit governance — board members should not use org resources for private benefit; Duty of loyalty",
     "During Larry's PMTNM presidency, the PMTNM website was used to promote NM School of Music. As incoming VP, Chip discovered this upon taking over web responsibilities. Larry's school derived promotional benefit from the nonprofit's digital presence.",
     "Nov 2024 'domain registrar and hosting' thread (Larry's web handoff to Chip); Website content during Larry's tenure; Chip's subsequent website documentation",
     "Direct private benefit from nonprofit's resources. Larry's school gained search presence and credibility from PMTNM's organizational identity.",
     "Documented via web handoff emails"),

    ("L-04", "Nov 5-6, 2025", "Conflict of Interest / Bias / Hostile Escalation", "Ethics + Procedural",
     "CRITICAL", "Larry Blind",
     "Robert's Rules of Order — Parliamentarian must be impartial; Due process rights of meeting participants; Fiduciary duty not to use authority for personal disputes",
     "ESCALATION SEQUENCE (Nov 5-7, 2025):\n\n(1) NOV 5, 6:18 PM — Chip sends professional governance email on existing 'Budget Committee REMINDER' thread. Documents serious compliance issues (5+ year filing gap) and respectfully asks Jeanne to retire with dignity. Explicitly states 'This is not an ultimatum.' References 'heavy and dark conversations' in context of scrutinizing documented failures — professional language for governance accountability.\n\n(2) NOV 5, 9:08 PM — DELIBERATE ESCALATION: Instead of responding on the Budget Committee thread, Larry creates a SEPARATE EMAIL THREAD with only Chip's name as subject line: 'Willis Glen Miller, III.' In this new thread (broadcast to full board: Sharon, Jeanne, Laura, Heather), Larry calls Chip's governance concerns 'offensive, disrespectful, and does not deserve a response.' This is not a reply to the budget/compliance issues — it's a personal attack designed to discredit Chip before the board meeting. The separate thread structure shows intentional escalation from procedural discussion to personal attack.\n\n(3) NOV 7, MORNING — Chip discovers Larry's hostile email the morning of the board meeting.\n\n(4) NOV 7, BOARD MEETING — Larry serves as Parliamentarian despite his demonstrated personal animus toward the VP. Minutes attribute a parliamentary 'time call' to him that does not appear in the audio recording. Larry uses his parliamentary authority to rule against Chip's no-confidence motion — the motion Chip raised after being personally attacked by Larry the night before.",
     "Nov 5, 6:18 PM email from Chip (on Budget Committee REMINDER thread): documents compliance issues, professional tone, explicit disclaimer 'This is not an ultimatum'; Nov 5, 9:08 PM email from Larry (NEW THREAD titled 'Willis Glen Miller, III', CC'd to Sharon/Jeanne/Laura/Heather): 'offensive, disrespectful, and does not deserve a response'; Nov 6 morning: Chip's discovery of hostile email sent night before meeting; Nov 7 Board Meeting Audio Recording showing no time call; Nov 7 Minutes recording fabricated 'time call'; INTERNAL_Minutes Review; Proposed Corrections to Minutes",
     "A board member personally attacked the VP in a separate, emphasized thread broadcast to the full board the night before a meeting. The next morning, that same person served as the neutral Parliamentarian and used parliamentary authority to rule against the VP's motion. The 'time call' recorded in the minutes does not appear in the audio — suggesting it was fabricated to justify the ruling. This is not impartiality; it's weaponization of a procedural role in service of a personal dispute.",
     "Timeline-verified, Audio-verified, Thread-structure-verified — Deliberate escalation from governance issue to personal attack to abuse of authority"),

    # ── MTNA ETHICS PROCESS ───────────────────────────────────────────────
    ("E-01", "Jul 27, 2024", "Ethics Process", "Ethics",
     "HIGH", "Larry Blind; Tatiana (NMSM); Brian Shepard (MTNA)",
     "MTNA Code of Ethics; Professional standards for teacher-parent communication during student transition",
     "Chip filed an ethical dispute mediation request with MTNA CEO Brian Shepard regarding Tatiana/NMSM blocking communication with Evan's mother during a student transition. Larry Blind allegedly spoke negatively to the parent about Chip, damaging Chip's professional relationship with the Menaul School. NMSM's employee handbook was invoked to block the MTNA ethics process from proceeding.",
     "Oct 12-Nov 16, 2023: Text message from Larry + email exchange (Gmail thread: 18bcf83c2e939e8c) — Larry's unsubstantiated accusation, parent email never produced, parent had called Chip (not vice versa), Larry refused to speak with Chip over weekend; Nov 16, 2023: Open communication request (Gmail thread: 18bd7f87933a6814) — Chip cited MTNA ethics code verbatim, offered professional mediation at neutral venue, offer never answered; Jul 27, 2024 email to Brian Shepard ('Ethical Dispute Mediation Request'); Aug 9 follow-up; Haewon thread Aug 31 (Chip: 'Under the New Mexico School of Music teaching contract, a breach of...'); Moving forward with PMTNM thread",
     "NMSM's employee handbook was used as a shield to block a legitimate MTNA ethics investigation. This is the same year Larry awarded Tatiana (his employee) the PMTNM Teacher of the Year award.",
     "Ethics complaint submitted — no action by MTNA"),

    ("E-02", "Nov 14, 2023", "Ethics Process", "Ethics",
     "HIGH", "Larry Blind; PMTNM Board; Gary Ingle (MTNA)",
     "MTNA Code of Ethics; Conference standards",
     "First ethics complaint filed with MTNA Executive Director Gary Ingle regarding the 2023 PMTNM conference — specifically Larry's use of the conference to feature NMSM students. Gary Ingle responded Dec 5, 2023 but took no action. Complaint forwarded to Cheryl Pachak-Brooks (ENMU board member) June 2024.",
     "Nov 14, 2023 email to Gary Ingle; Dec 5, 2023 response from Ingle; Feb 2024 and Jun 2024 follow-up forwards",
     "Pattern of MTNA declining to hold affiliates accountable. Both complaints — Nov 2023 and Jul 2024 — resulted in no action despite documented violations.",
     "No action taken — Pattern established"),

    ("E-03", "Sep 2025; Dec 2025; Jan 2026", "Ethics Process", "Ethics",
     "HIGH", "Brian Shepard (MTNA); Jeanne Grealish",
     "MTNA affiliate compliance standards; MTNA's responsibility to member affiliates",
     "Chip files third complaint to Brian Shepard (Sep 23, 2025) about Grealish, attaching a full ZIP file of documentation. Shepard confirms receipt. Chip files formal PMTNM affiliate compliance complaint Dec 23, 2025. Shepard responds Jan 6-7, 2026 with platitudes and defers entirely to affiliate self-governance. Three separate complaints to MTNA national — zero enforcement action.",
     "Sep 23, 2025 email to Shepard with ZIP documentation; Dec 23, 2025 formal complaint; Jan 6 Shepard receipt confirmation; Jan 7 Shepard response (defers to affiliate)",
     "MTNA CEO's pattern of acknowledging complaints and taking no action suggests either institutional capture or policy of non-interference that effectively shields bad actors.",
     "Three complaints filed — zero action"),

    ("E-04", "Oct 2023 - Jul 2024", "Organizational Obstruction of Ethics Process", "Ethics + Governance",
     "HIGH", "Lawrence Blind (NMSM authority); Tatiana Vetrinskaya (NMSM co-founder)",
     "MTNA Code of Ethics § Commitment to Colleagues (teacher participation in student transitions); Professional standards for teacher-parent communication; Fiduciary duty not to use organizational position to block accountability",
     "When Chip Miller (piano teacher, MTNA member) withdrew from New Mexico School of Music in October 2023, a parent (Toni Morgan) was upset about the transition and lack of communication. Miller attempted to facilitate an open conversation between the parent and NMSM to resolve confusion — consistent with MTNA Code of Ethics: 'The teacher shall participate in the student's change of teachers with as much communication as possible between parties.'\n\nLarry Blind (NMSM authority) blocked this communication via text message: 'Chip, please do not contact any of your former students from our school by phone or email. This is not appropriate and I have received an email from a parent who was not very happy about your call.'\n\nTatiana Vetrinskaya (NMSM co-founder/owner) reinforced the block, claiming exclusive responsibility: 'as the school owner, I am responsible for any concerns or issues our teachers or clients may express. From now on, please communicate with me directly.'\n\nMiller escalated the matter to MTNA Code of Ethics compliance, filing an Ethical Dispute Mediation Request with MTNA CEO Brian Shepard (Jul 27, 2024). Shepard reviewed materials but ultimately could not process the complaint because NMSM's handbook language created a structural conflict with MTNA's Code of Ethics: the school claimed exclusive control over all student/parent communication, preventing the teacher from fulfilling their MTNA ethical duty to participate in transitions.\n\nThis demonstrates Larry Blind's pattern of using organizational structure and handbook language to create shields against professional accountability. The mechanism is systematic: design bylaws/handbooks to conflict with external ethical standards, then invoke those conflicting rules to block oversight.",
     "Oct 12, 2023, 12:57 PM text message from Larry Blind to Chip Miller: 'Chip, please do not contact any of your former students from our school by phone or email. This is not appropriate and I have received an email from a parent who was not very happy about your call'; Nov 14, 2023, 5:43 PM email from Tatiana Vetrinskaya: 'as the school owner, I am responsible for any concerns or issues our teachers or clients may express. From now on, please communicate with me directly'; Nov 14, 2023, 7:55 PM email from Chip Miller explaining ethical violation (MTNA Code violation, poor communication from school, Larry's inappropriate reprimand); Jul 27, 2024: Ethical Dispute Mediation Request filed by Chip Miller to Brian Shepard (bshepard@mtna.org); MTNA's subsequent inability to process complaint due to NMSM handbook conflicts with MTNA Code of Ethics; Email thread 190f4f767cafa7ff (Gmail search result)",
     "Pattern across two organizations: At PMTNM, Larry blocks anonymous voting systems and uses procedural rules to prevent board oversight. At NMSM, Larry blocks professional ethics compliance using school handbook language that conflicts with MTNA standards. In both cases, he systematically designs organizational structures to insulate himself from accountability — creating rules that prevent external standards from being applied. This is not coincidental governance variation; it is deliberate structural obstruction.",
     "Text + email verified — Pattern consistent with L-04 (bias as Parliamentarian) and L-07 (undisclosed roles to control processes) — Shows systematic use of organizational position to block all forms of oversight and accountability"),

    # ── BOARD MINUTES IRREGULARITIES ──────────────────────────────────────
    ("M-01", "Nov 7, 2025", "Minutes Falsification", "Legal",
     "CRITICAL", "Jeanne Grealish (minutes author); Sharon Kunitz (President, meeting chair)",
     "NM Nonprofit Corporation Act — accurate records requirement; Robert's Rules on minute accuracy; Potential falsification of official records",
     "IRREGULARITY 1: Minutes state 'a 10 minute limit for each speaker.' Recording and Sharon's own pre-meeting email (7:07 AM Nov 7) confirm the limit was 15 minutes. The president's own written email contradicts her secretary's minutes.",
     "Sharon Kunitz email Nov 7, 2025 at 7:07 AM ('Additional RRoo wisdom!'): '15 minutes to speak'; Nov 7 Board Meeting Audio Recording at ~2:41; Proposed Corrections to Minutes (Apr 10, 2026)",
     "Minutes misrepresent an established rule — documented in the president's own email hours before the meeting. Creates false record that VP exceeded time limits.",
     "Audio-verified + Email-verified"),

    ("M-02", "Nov 7, 2025", "Minutes Falsification", "Legal",
     "CRITICAL", "Jeanne Grealish",
     "NM Nonprofit Corporation Act; Accurate records; Potential fraud",
     "IRREGULARITY 2: Minutes state financial report 'was approved as presented.' No motion, second, or vote appears anywhere in the audio recording. The only financial motion made was about CPA-designated accounts — which minutes appear to conflate with general financial approval.",
     "Nov 7 Board Meeting Audio Recording; Nov 7 Transcript; INTERNAL_Minutes Review (audio-timestamped); Proposed Corrections to Minutes",
     "False approval of financial report creates appearance of board oversight where none occurred. Potentially used to legitimize financial decisions that were never actually voted on.",
     "Audio-verified — Most serious minutes issue"),

    ("M-03", "Nov 7, 2025", "Minutes Falsification", "Legal",
     "CRITICAL", "Jeanne Grealish; Larry Blind (named as actor)",
     "Robert's Rules — parliamentary procedure; Accurate records",
     "IRREGULARITY 3: Minutes state 'time was called by the Parliamentarian.' No such event appears in the audio recording. At 49:13, President adjourns meeting. At 48:22, Parliamentarian (Larry Blind) says 'we can't discuss it unless somebody seconds it.' No time call occurs.",
     "Nov 7 Board Meeting Audio Recording at 48:22 and 49:13; INTERNAL_Minutes Review; Proposed Corrections to Minutes",
     "Fabricated parliamentary event used to justify cutting off VP's presentation. Names Larry Blind as performing an action he did not perform.",
     "Audio-verified — Strong contradiction"),

    ("M-04", "Nov 7, 2025", "Minutes Falsification + Document Suppression", "Legal",
     "CRITICAL", "Jeanne Grealish; Sharon Kunitz; Larry Blind",
     "Accurate records requirement; Robert's Rules — reports submitted to the secretary become part of the record; VP's right to submit reports; NM Nonprofit Act — records of officer reports",
     "IRREGULARITY 4: Minutes state VP report 'had not been submitted with other Board reports.' The audio recording proves the opposite — the report was physically handed to both Sharon Kunitz AND Jeanne Grealish before the meeting.\n\nVERBATIM FROM NOV 7 RECORDING:\n- Chip at 32:46: 'Yes, on paper, handed to Sharon and handed to Jean.'\n- Grealish at 33:05: 'the lensing paper that you gave us with the attorney\'s office' — Grealish IDENTIFIED the document but then said (33:18): 'It cannot be board reports, because I don\'t have anything to include.'\n- Chip at 33:27: 'As I said, a copy was given to Sharon and Jean.'\n- Larry at 33:34: 'If we have a paper copy, I can run it off on our machines.'\n- Larry ran physical copies on the copy machine during the meeting — proving the document existed and was in their possession.\n- Larry at 34:11, after returning from copy machine: 'it does not say that it is from you' [questioning authorship as a delay tactic]\n- Chip at 34:49: 'The first sentence is this end of year report provides the board of directors with a comprehensive assessment of the Vice President\'s activities.'\n\nSEQUENCE OF SUPPRESSION: Report physically handed to Sharon + Jean → Both deny having it → Report surfaces during meeting → Larry copies it → Larry questions authorship → Sharon asks Chip to summarize verbally 'until Larry returns' → Board reads the report selectively → Sharon moves to adjourn without acting on it → Minutes record VP report was 'not submitted.'\n\nGrealish\'s statement 'It cannot be board reports, because I don\'t have anything to include' is admission that she controls what becomes an official board record. By claiming not to 'have anything to include,' she unilaterally excluded Chip\'s submitted report from the record — then the minutes recorded her exclusion as fact.",
     "VP_Report_Submitted_to_President_and_EST.pdf (Google Drive ID: 1d-vg5v87N7DwF7EUisJ6M-HYiYrbuC3Z); Nov 7 Audio Recording at 32:46 (Chip: 'handed to Sharon and handed to Jean'); 33:05 (Grealish identifies the document); 33:18 (Grealish: 'It cannot be board reports, because I don\'t have anything to include'); 33:34 (Larry: 'I can run it off on our machines'); 34:11 (Larry copy machine trip + authorship challenge); Nov 7 Minutes stating report 'had not been submitted'",
     "The minutes record as fact something that is directly contradicted by audio: the report was physically present, identified by Grealish, copied by Larry, and read aloud by Larry. Grealish\'s statement that she controls what becomes an official board report is the mechanism of falsification. The copy machine trip proves the document existed. The minutes\'s false claim about non-submission is the cover-up.",
     "AUDIO-VERIFIED — report physically present, identified, copied during meeting, then recorded as 'not submitted'"),

    ("M-05", "Nov 7, 2025", "Procedural Violation — VP Excluded from Board Packet", "Governance",
     "HIGH", "Sharon Kunitz; Jeanne Grealish",
     "VP's right to receive board meeting materials; Basic procedural fairness; Equal treatment of officers",
     "VP Miller stated on recording at 32:28: 'I have no packet.' This was not disputed by any other attendee. Larry Blind, after retrieving Chip\'s VP report from the copy machine, said at 34:11: 'it does not say that it is from you' — implying he was working from a document not in the standard packet. The board packet had been distributed to other members in advance but Chip received nothing.\n\nAt 42:22, after Sharon attempted to adjourn: Chip: 'I need to note that there are items on the list that were sent to you that have not been covered.' At 42:43, an unidentified voice: 'You didn\'t send that.' At 42:45, Chip: 'I sent you an email, Sharon, it was very specific with board agenda items.' At 42:52, Sharon herself: 'two months ago, March 17, thank you' — confirming receipt and inadvertently acknowledging she had held the items for two months without placing them on the agenda.\n\nMINUTES: Contain no reference to VP\'s lack of packet, no reference to the agenda items dispute, no reference to Sharon\'s adjournment attempt or Chip\'s objection. 'Old Business: None. New Business: None.'",
     "Nov 7 Audio Recording at 32:28 (Chip: 'I have no packet'); 42:22 (items not covered); 42:45 (email confirmation); 42:52 (Sharon: 'two months ago, March 17, thank you'); Oct 14, 2025 VP Agenda Request email (the items submitted); Nov 7 Board Meeting minutes ('Old Business: None'); Proposed Corrections to Minutes #8 and #9",
     "VP was excluded from the pre-meeting board packet, had his submitted agenda items ignored, and when he objected to adjournment, was told by one voice that he\'d never sent items — immediately contradicted by the President herself, who confirmed receipt from two months prior. The minutes erase the entire dispute.",
     "Audio-verified + Sharon\'s own self-contradiction on record"),

    # ── SHARON KUNITZ — CONFLICT OF INTEREST ──────────────────────────────
    ("S-01", "Ongoing 2025", "Conflict of Interest", "Ethics",
     "HIGH", "Sharon Kunitz; Jeanne Grealish",
     "Nonprofit governance — board members must be free of conflicts; Fiduciary duty",
     "Sharon Kunitz receives scholarship money from PMTNM. Jeanne Grealish has sole sign-off authority on scholarship disbursements. Sharon consistently backs Jeanne on all governance disputes, financial access questions, and VP complaints — creating the appearance of a quid pro quo relationship.",
     "DocsFromJeanne financial records (scholarship disbursements); Spring 2025 Board Meeting records; Pattern of Sharon defending Grealish across all email threads",
     "If Sharon's scholarship payments are controlled by Grealish, Sharon has a financial interest in maintaining Grealish's power. This creates an undisclosed conflict of interest on all votes where Sharon sides with Grealish.",
     "To be confirmed from financial records"),

    ("S-02", "Mar 17, 2025", "Governance Violation", "Ethics",
     "HIGH", "Sharon Kunitz",
     "VP's right to financial records; PMTNM Handbook on VP duties",
     "When Chip requested financial records, Sharon reinforced Grealish's refusal by telling Chip his duties were 'collecting funds for the MTNA Foundation and recruiting members' — a reductive mischaracterization designed to strip the VP of governance authority.",
     "Email Mar 17, 2025 (Sharon → Chip, responding to financial request); VP's documented handbook duties including Budget Committee membership",
     "President actively undermining VP's legal right to financial records. Combined with Grealish's refusal, establishes coordinated obstruction.",
     "Documented — Key supporting violation"),

    ("S-03", "Nov 4–7, 2025", "Broken Commitment + Attempted Wrongful Adjournment", "Legal",
     "CRITICAL", "Sharon Kunitz",
     "Good faith obligations; Promissory estoppel; Robert's Rules §23 (adjournment — cannot adjourn when business is pending); Due process in nonprofit governance",
     "TWO SEPARATE VIOLATIONS, FOUR DAYS APART, ON AUDIO RECORD:\n\n(1) NOV 4, 2025 at 1:14:51 (recorded): After hearing VP's full report and Kathy Black's legal recommendations, Sharon Kunitz stated: 'I'll put them on the agenda.' Attorney Kathy Black immediately confirmed on the recording: 'She's going to put them on the agenda for Friday morning.' This commitment was made in the presence of legal counsel. Kathy Black's same-day email (2:54 PM) further confirmed: 'Sharon agreed to tee up certain agenda items.'\n\n(2) NOV 7, 2025 at approximately 41:30 (recorded): Without addressing any of Chip's submitted agenda items, Sharon announced: 'I did not have agenda items. So meeting is adjourned.' Chip immediately objected: 'I need to note that there are items on the list that were sent to you that have not been covered.' Sharon initially deflected, then accidentally confirmed receipt: 'two months ago, March 17, thank you' — acknowledging she had received the items while also revealing she'd held them for two months without acting.\n\nADDITIONAL: An unidentified voice at 42:43 said 'You didn't send that' about Chip's agenda submission. Sharon's own response immediately contradicted this: 'two months ago, March 17, thank you.' Sharon herself confirmed receipt while the meeting was trying to deny it.\n\nThe Nov 7 minutes record 'Old Business: None. New Business: None.' — erasing any trace of the dispute.",
     "Oct 14, 2025 VP Agenda Request email to Sharon (subject: ITEMS FOR BOARD AGENDA Nov 7-8, Gmail thread: 199e36adc51ecfde) — five formal items invoking Handbook Art. VIII §5 and three NM statutes; Sharon's reply Oct 14 at 16:08: 'As I have said, there is more in my life than PMTNM... Will have more time for this tomorrow morning.' — no substantive engagement with any item, no follow-up; Nov 4, 2025 Audio Recording at 1:14:51 (Sharon: 'I'll put them on the agenda'); Nov 4, 2025 Audio Recording at 1:14:53 (Kathy Black: 'She's going to put them on the agenda'); Kathy Black email Nov 4 at 2:54 PM confirming agenda commitment; Nov 7, 2025 Audio Recording at ~41:30 (Sharon: 'I did not have agenda items. So meeting is adjourned.'); Nov 7 Audio Recording at ~42:52 (Sharon: 'two months ago, March 17, thank you' — confirming receipt of Chip's earlier communication); Nov 7 Minutes: 'Old Business: None. New Business: None.'",
     "Sharon made a witnessed commitment before an attorney, confirmed immediately by that attorney on recording, then reversed it four days later — attempting to adjourn the meeting without addressing a single submitted agenda item. When challenged, she accidentally confirmed having received the items two months earlier. This is the clearest single sequence demonstrating intentional governance exclusion: commitment, reversal, concealment, then accidental self-contradiction.",
     "DUAL AUDIO VERIFIED (Nov 4 + Nov 7) + Attorney email + Sharon's own self-contradiction on record"),

    # ── RECORDS DISTRIBUTION FAILURES ─────────────────────────────────────
    ("R-01", "Jan 2025 → Jul 2025", "Records Access", "Legal",
     "HIGH", "Jeanne Grealish",
     "PMTNM Handbook: Handbook should be furnished to each member; NMSA §53-8-26",
     "Chip requested the PMTNM Handbook in January 2025. Grealish falsely claimed it was available online (it was not — offline for 2+ years). Chip did not receive a copy until July 13, 2025 — 6 months later.",
     "Jan 2025 Handbook thread; Jul 13, 2025 email distributing handbook; Grealish's false claim of online availability",
     "6-month obstruction of access to governing documents is itself a governance violation. Compound with financial records refusal.",
     "Documented"),

    ("R-04", "Date of distribution unknown (referenced in 'PMTNM. LIST AND EMAIL ADDRESSES' thread)", "Inaccurate Credential Documentation", "Records + Governance",
     "MEDIUM", "Jeanne Grealish",
     "Nonprofit record-keeping standards; Accuracy of member status documentation; Fiduciary duty to maintain accurate organizational records; MTNA credential verification",
     "PMTNM's official internal member list ('PMTNM. LIST AND EMAIL ADDRESSES') lists Sharon Kunitz as 'NCTM' (National Certified Teacher of Music). Official MTNA data (MTNAData.csv, dated 2024) shows Sharon Kunitz with a blank DESIGNATION field — no NCTM credential listed. By contrast, verified NCTMs in the same MTNA dataset (Jeanne Grealish, Cheryl Pachak-Brooks, and others) all display 'NCTM' in the DESIGNATION field. NCTM is an earned honor from MTNA requiring ongoing fees and maintenance. Sharon either never attained the credential, allowed it to lapse, or failed to renew. Grealish, as Executive Secretary-Treasurer responsible for accurate member records, circulated an official organizational document misrepresenting Sharon's status. This is particularly problematic because: (1) it damages organizational credibility with external parties, (2) it reflects the pattern of inaccurate record-keeping by Grealish (R-01, R-02, R-03, M-04), and (3) it involves Sharon, who has been Grealish's consistent supporter on the board despite receiving scholarship funds controlled by Grealish (S-01).",
     "Official MTNA data (MTNAData.csv): Sharon Kunitz (row 40) shows blank DESIGNATION field vs. Jeanne Grealish (row 30) and Cheryl Pachak-Brooks (row 52) both showing 'NCTM' in DESIGNATION field; PMTNM internal member list email thread ('PMTNM. LIST AND EMAIL ADDRESSES'): lists Sharon as NCTM; Email attachment: 'Email addresses (2).xlsx' showing same false NCTM designation for Sharon",
     "False representation of member credentials in official organizational documentation. Pattern of inaccurate EST record-keeping. Damages organizational credibility. Misrepresentation could affect external perceptions of PMTNM's accuracy and governance.",
     "Documented — MTNA data vs. PMTNM internal list verified"),

    ("R-02", "Nov 7, 2025 → Mar 9, 2026", "Records Distribution", "Governance",
     "HIGH", "Jeanne Grealish",
     "Standard nonprofit governance — minutes should be distributed within 30-60 days; PMTNM handbook on minutes distribution",
     "November 7, 2025 board meeting minutes were dated January 1, 2026 and not distributed until March 9, 2026 — four months after the meeting. Distribution came after the petition was already filed (Feb 2026) and after PMTNM's answer was filed (Mar 5, 2026).",
     "Jan 1, 2026 date on minutes; Mar 9, 2026 distribution email from Grealish; Petition filed Jan 27, 2026; PMTNM Answer filed Mar 5, 2026",
     "4-month delay in distributing minutes is itself a governance failure. Timing of distribution (after legal proceedings filed) raises questions about deliberate delay.",
     "Documented — Suspicious timing"),

    ("R-03", "Apr 25, 2025 → Nov 7, 2025", "Records Distribution", "Governance",
     "MEDIUM", "Jeanne Grealish",
     "Standard practice — spring board meeting minutes should be distributed before fall meeting",
     "As of Oct 9, 2025, Grealish stated spring minutes had not been distributed and would be 'compiled and sent with Nov first' board reports. Spring minutes (April 2025) were not circulated until November — a 7-month delay. This was noted in Chip's Oct 14 agenda request email.",
     "Oct 9, 2025 Grealish email re: minutes distribution; Oct 14 VP Agenda Request email documenting this; Nov 7 Board Meeting (April minutes addressed at start of meeting)",
     "Pattern of chronic delays in distributing meeting records. Creates information asymmetry between Grealish-aligned insiders and officers like Chip who have no independent access.",
     "Documented"),

    # ── LARRY BLIND — TOTY MANIPULATION (DETAILED) ────────────────────────
    ("L-05", "Oct 2023", "TOTY Manipulation / Boss-as-Nominee", "Ethics + Governance",
     "CRITICAL", "Larry Blind (NMSM employee); Tatiana Vetrinskaya (NMSM co-founder/owner)",
     "MTNA Code of Ethics §3 (conflicts of interest); Nonprofit election integrity; Robert's Rules on impartiality; Fiduciary duty of loyalty",
     "2023 TOTY had two nominees: Terri Reck (nomination written by Jackie Zander-Wall, AMTA president) vs. Tatiana Vetrinskaya (nomination written by Larry Blind, his BOSS and NMSM co-founder/owner). Larry administered the vote from his business email (lblind@nmschoolofmusic.com) and explicitly stated in his voting email: 'I decided to extend the nomination period' — a unilateral extension with no board approval. Larry's employment at NMSM depends on Tatiana's goodwill, making impartial administration impossible. Chip's side-by-side analysis: Terri Reck's bio = 'dense list of information highlighting work done in the PMTNM community this past year.' Tatiana's bio (written by Larry) = 'storybook language throughout that causes emotional reactions, often unrelated to the question at hand. Highlights work done in NMSM.' Haewon Yang admitted voting for 'the person whose bio looked the best.' Tatiana won. No conflict of interest was disclosed by Larry.\n\nVERBATIM FROM TATIANA'S ACTUAL SUBMISSION (submitted by lawrenceblind@msn.com, signed 9-30-23): Community service section = entirely NMSM activities. Closing line: 'business acumen... it is what I know she is most proud of.' [Note: 'business acumen' is not a PMTNM TOTY criterion.] Future-tense item listed under achievements: 'will present a workshop at 2023 PMTNM Conference' — event had not yet occurred at time of submission. Tatiana's listed email: tatiana@nmschoolofmusic.com (NMSM institutional address).\n\nVERBATIM FROM TERRI RECK'S ACTUAL SUBMISSION (submitted by zanderwall@gmail.com / Jacqueline Zander-Wall, signed 9/30/23): 47 years of teaching (1976–2023). Dalcroze certification across 6 institutions. 5 AMTA committee leadership roles. Harpsichord performances with SF Symphony, NM Symphony, NM Philharmonic. Students enrolled in PMTNM/MTNA programs. Music Bowl Chair for decades. Jewelry donated to PMTNM silent auctions.",
     "Oct 4, 2023 email from lblind@nmschoolofmusic.com TO lblind@nmschoolofmusic.com (BCC to board): Larry explicitly states 'I decided to extend the nomination period'; Two TOTY PDFs sent with vote solicitation: 'Terri Reck.pdf' and 'Tatiana Vetrinskaya.pdf' (Gmail thread ID: 19239a27e03b7010); PHYSICAL PDFs READ — Tatiana's PDF: submitted by lawrenceblind@msn.com, Tatiana email tatiana@nmschoolofmusic.com, closing 'business acumen — it is what I know she is most proud of,' future-tense 'will present a workshop at 2023 PMTNM Conference,' all community service through NMSM; Terri's PDF: submitted by zanderwall@gmail.com, 47 years teaching, Dalcroze certification, 5 AMTA committee roles, harpsichord with major orchestras, decades as Music Bowl Chair, jewelry to PMTNM silent auctions; Haewon thread Sep 27-28, 2024: 'last year I voted for Tatyana because I had no idea who these people were and just voted for the person whose bio looked the best'; Chip to Jackie Sep 29, 2024: 'The other submission paints a picture of a larger timeframe and uses storybook language throughout that causes emotional reactions, and are often unrelated to the question at hand. It highlights work done in NMSM'; Jacqueline Zander-Wall response: 'I am not contesting the decision made at all. I think one needs to move on.'",
     "Employer wrote a polished 'storybook' bio for his employee vs. an independent teacher's factual PMTNM community service record. The winning submission listed a future event (not yet a completed achievement), described 'business acumen' (not a TOTY criterion), and used the nominee's NMSM institutional email — all pointing to an NMSM marketing document rather than a community service record. The losing submission documented 47 years of verifiable community service across PMTNM, AMTA, and major NM orchestras. Vote administered by the conflicted party from NMSM business email. Unilateral deadline extension. Haewon's admission confirms votes cast on presentation quality, not merit. Jackie Zander-Wall (AMTA president) declined to contest — the quiet resignation of a community teacher facing institutional power.",
     "Email evidence + Haewon on-record admission + Chip comparative analysis (verbatim) + PHYSICAL PDFs READ AND ANALYZED"),

    ("L-06", "Ongoing 2023–2025", "NMSM Business Promotion via PMTNM", "Ethics + Financial",
     "HIGH", "Larry Blind",
     "MTNA Code of Ethics: 'Represent our art honestly'; IRS private benefit prohibition; Nonprofit duty of loyalty",
     "Larry's NMSM director biography uses systematically misleading language: 'students have an opportunity to perform internationally' (they pay to attend European recitals — this is not performance, it is tourism); 'returned for an encore performance' at Carnegie Hall (how he accessed the hall is unsubstantiated — the concern is the misleading framing). NMSM employs uncredentialed faculty (flute majors teaching piano, teachers in first jobs). Larry used PMTNM's conference platform and handbook to promote NMSM, diluting PMTNM value to independent member teachers whose market is directly harmed by NMSM's growth.",
     "NMSM director biography (nmschoolofmusic.com/director---lawrence-blind — live web evidence); 2023 PMTNM conference program featuring NMSM students; VP Report documenting biography misrepresentation; PMTNM member impact analysis",
     "Using nonprofit organizational credibility and platform to market a competing commercial music school. Misleading language in NMSM materials violates MTNA's core principle of honest representation. NMSM's 700-student enrollment at $70/hr directly harms independent PMTNM members' livelihoods.",
     "Web evidence + conference records + VP Report"),

    # ── FINANCIAL — VP REPORT ADDITIONAL EVIDENCE ─────────────────────────
    ("F-10", "Nov 2025", "Financial Self-Dealing", "Ethics + Financial",
     "HIGH", "Jeanne Grealish",
     "IRS private benefit prohibition; Fiduciary duty; PMTNM budget oversight",
     "Storage unit operating expense documented as $4,600 actual vs. $3,000 budgeted — a 53% overrun. Even Larry Blind, who has a financial relationship with Grealish, raised the discrepancy in a Nov 2 email asking about the expense. Grealish had already admitted (before attorney Kathy Black) that she stores personal furniture in the PMTNM storage unit.",
     "Nov 2, 2025 Larry Blind email raising $4,600 storage question; Nov 4 meeting transcript (Grealish personal furniture admission before Kathy Black); DocsFromJeanne budget documents showing $3,000 budget vs. $4,600 actual",
     "53% budget overrun on personal-use storage. Even the past president questioned it in writing. Attorney-witnessed admission of personal property in PMTNM storage creates clear private-benefit violation.",
     "Attorney-witnessed + Larry Blind email confirmation"),

    # ── ETHICS PROCESS — TOTY PROCEDURAL VIOLATION ────────────────────────
    ("E-04", "Oct 2023", "TOTY Procedural Violation", "Ethics + Governance",
     "HIGH", "Larry Blind; Tatiana Vetrinskaya (NMSM)",
     "MTNA Code of Ethics — conflicts must be disclosed; PMTNM TOTY award criteria (Criterion B3: 'outstanding professional relationship with students, fellow music teachers, and community leaders'; B4: 'consistent record of high standards of teaching and promoting music teaching in New Mexico')",
     "The TOTY award criteria require demonstrated community service within New Mexico's music teaching community. Criterion-by-criterion analysis of the actual PDFs:\n\nTERRI RECK (loser): B1 ✓ — 47 years teaching, Dalcroze certification. B2 ✓ — 5 AMTA committee leadership roles, SF Symphony, NM Symphony, NM Philharmonic harpsichordist. B3 ✓ — Students in PMTNM/MTNA programs, Music Bowl Chair decades, jewelry donations to PMTNM silent auctions. B4 ✓ — Entire career in NM; engaged with PMTNM community across multiple decades.\n\nTATIANA VETRINSKAYA (winner, written by her employer Larry Blind): B1 ✓ — teaching background listed. B2 — ambiguous; achievements framed around NMSM institutional activity. B3 ✗ — community service section entirely NMSM activities, not PMTNM/NM music teaching community. B4 ✗ (future tense disqualifier) — submission lists 'will present a workshop at 2023 PMTNM Conference' — an event that had NOT YET OCCURRED when submission was filed on 9-30-23. Closing attribute 'business acumen' is not a TOTY criterion. No conflict of interest disclosed.",
     "PHYSICAL PDFs read and analyzed: 'Terri Reck.pdf' (submitted zanderwall@gmail.com, signed 9/30/23) and 'Tatiana Vetrinskaya.pdf' (submitted lawrenceblind@msn.com, signed 9-30-23); Larry's Oct 4, 2023 voting email with award criteria (Gmail thread ID: 19239a27e03b7010): B3 and B4 require PMTNM community service; Chip to Jackie Sep 29 (verbatim comparison): 'The other submission... highlights work done in NMSM' vs. 'Your submission... highlights work done in the PMTNM community this past year'; Haewon admission that she voted based on bio appearance; Tatiana's future-tense 'will present a workshop' listed as an achievement",
     "Award designed to honor PMTNM community service given to NMSM institutional promotion. The winning bio fails B3 entirely and lists a future event as an achievement. The losing bio meets all four criteria with verifiable, decades-long records. No conflict disclosure made by the employer who wrote the winning nomination and administered the vote.",
     "PHYSICAL PDFs analyzed + verbatim comparative analysis in email record"),

    # ── SHARON KUNITZ — FULL DOCUMENTED PATTERN ───────────────────────────
    ("S-04", "Mar 2025 – Feb 2026", "Governance Obstruction", "Ethics + Legal",
     "HIGH", "Sharon Kunitz",
     "VP's statutory right to financial information (NMSA §53-8-26); PMTNM Handbook on VP Budget Committee membership; Presidential duty of good faith",
     "For 11+ months, Sharon used her position as President to block Chip's financial access: (1) Mar 2025 — reinforced Grealish's records refusal, falsely characterizing VP's duties; (2) Oct 2025 — dismissed budget committee requests by deferring to Grealish; (3) Oct 2025 — told Chip that scholarship coordination was 'not your territory'; (4) Nov 4 — committed to place agenda items before board, then broke that commitment at Nov 7.",
     "Mar 17, 2025 email (Sharon reinforces Grealish refusal); Oct 2 email (budget committee deferral); Oct 14 agenda request reply (dismissive); Nov 4 audio recording (commitment at 1:14:51); Kathy Black email Nov 4 at 2:54 PM (confirms commitment); Nov 7 meeting (commitment broken)",
     "Sustained 11-month pattern of presidential obstruction of VP's statutory inspection rights. Pattern of coordination with Grealish to maintain financial information asymmetry.",
     "11-month documented pattern — four distinct incidents"),

    ("S-05", "Sep – Oct 2025", "Failure to Act on Known Misconduct", "Ethics",
     "HIGH", "Sharon Kunitz",
     "MTNA Code of Ethics — board members must not tolerate known misconduct; Presidential responsibility for board conduct standards; Duty of candor",
     "Sep 11, 2025: Jeanne Grealish responded to Chip's MTNA Foundation newsletter: 'Boring and irrelevant to most of us.' Sep 12, 12:44 AM: Chip forwarded Jeanne's response to Jeanne, Brian Shepard, AND Sharon Kunitz (slkunitz@aol.com) asking 'Hi Jeanne, can you clarify?' Sharon received this email — and said nothing. No defense of Chip's work, no pushback on Jeanne's dismissive characterization. Oct 2, 2025 — 21 days later — Sharon sent Chip a separate email: 'Great article on Foundation! Beautifully written and presented.' (CCed to Jeanne, Larry, Laura Spitzer). Sharon knew the article had merit. She praised it privately. She chose not to say so when Jeanne publicly dismissed it to Chip's face.",
     "Sep 12, 2025 12:44 AM email: Chip → Jeanne + bshepard@mtna.org + slkunitz@aol.com: 'Hi Jeanne, can you clarify?' (Gmail thread ID: 1993b54101446b14); Oct 2, 2025 Sharon → Chip (CC Jeanne, Larry, Laura): 'Great article on Foundation! Beautifully written and presented.' (Gmail thread ID: 199a581f54374d6f); No Sharon response in the Sep 12 thread",
     "Sharon's silence in the moment — combined with her private praise 21 days later — proves she recognized Jeanne's conduct was wrong and chose not to intervene. As President, she had the authority and responsibility to address Grealish's conduct toward a fellow officer.",
     "Two emails, 21 days apart — documentary contrast"),

    # ── VP REPORT FINDINGS — GREALISH ADMISSIONS ON RECORD ───────────────
    ("G-01", "Oct 7, 2025", "Legal Compliance Repudiation", "Legal + Ethics",
     "CRITICAL", "Jeanne Grealish",
     "IRS 501(c)(3) obligations; NM Nonprofit Corporation Act; Fiduciary duty to act within statutory constraints; PMTNM Handbook (Bylaws) provisions for budget committee and fiscal governance",
     "On Oct 7, 2025, Jeanne Grealish sent an email to Chip, Sharon, Laura, Larry, and Heather Nasi. The FULL CONTEXT: Chip had asked about the budget committee meeting required by the PMTNM Handbook. Grealish responded by dismissing the handbook's budget committee provisions as 'grossly out of date,' then wrote: 'Thanks heaven, PMTNM does not function under any legal mandates!' This statement was NOT an offhand philosophical remark — it was a deliberate, specific rejection of the handbook governance requirements Chip was invoking to justify a budget committee meeting. Grealish also stated during this period: conference venue research was 'a waste of everyone's time and a senseless cluttering up of mail boxes'; the website was 'a big waste of time'; 'Please, no more useless emails' (re: venue research). The statement is additionally documented in the VP Report submitted to the President and EST (Section 3.3).",
     "Oct 7, 2025 email from Grealish → Chip + Sharon + Laura + Larry + Heather Nasi (Gmail thread ID: 199c0d60237afa6c): VERBATIM: 'As for any statements in the Handbook regarding the budget, committee, dates, etc. these would be grossly out of date since the Handbook itself is no longer current and requires serious revision. Thanks heaven, PMTNM does not function under any legal mandates!'; VP_Report_Submitted_to_President_and_EST.pdf Section 3.3 (Drive ID: 1d-vg5v87N7DwF7EUisJ6M-HYiYrbuC3Z) independently documenting the same statement",
     "EST explicitly disclaiming legal obligations on the record — and doing so SPECIFICALLY to shut down VP's invocation of handbook governance requirements for the budget committee. The person controlling all of PMTNM's finances and records publicly stated the organization operates without legal mandates in response to a compliance question. This is the most unambiguous single statement supporting the petition and a future AG complaint.",
     "VERBATIM email + VP Report — dual source verification of same statement"),

    ("F-11", "2025", "TOTY Process Interference — Terri Reck Pattern", "Ethics + Governance",
     "HIGH", "Jeanne Grealish",
     "Nonprofit election integrity; MTNA Code of Ethics — fair processes; PMTNM TOTY procedural standards",
     "CRITICAL PATTERN: Terri Reck was ALSO the 2023 TOTY nominee who lost to Tatiana Vetrinskaya when Larry administered the biased process (L-05, E-04). In 2025, Grealish made three separate attempts to nullify Reck's nomination again, stating: 'I believe the reason Terri Reck was not nominated last time was she has not contributed very much to community service.' This claim was immediately contradicted by other board members who seconded Reck's nomination. Grealish's 'last time' reference is itself an admission — she knows Terri was nominated before and is aware of the prior cycle. Terri Reck — an independent community teacher with 47 years of teaching, 5 AMTA committee roles, and decades as Music Bowl Chair (per her own 2023 TOTY submission) — has been the target of TOTY suppression in back-to-back cycles: 2023 (Larry's biased process blocked her) and 2025 (Grealish's three direct nullification attempts).",
     "VP_Report_Submitted_to_President_and_EST.pdf — 2025 TOTY section (Google Drive ID: 1d-vg5v87N7DwF7EUisJ6M-HYiYrbuC3Z): Grealish: 'I believe the reason Terri Reck was not nominated last time was she has not contributed very much to community service'; Board members seconding Reck contradicting Grealish; 2023 TOTY PDF 'Terri Reck.pdf' (submitted zanderwall@gmail.com, Gmail thread ID: 19239a27e03b7010) documents Reck's actual community record: 47 years teaching, Dalcroze certification, SF/NM Symphony harpsichordist, Music Bowl Chair, PMTNM silent auction donations",
     "The same independent community teacher — with a documented record that exceeds TOTY criteria — has been suppressed in consecutive cycles by different insiders: Larry in 2023, Grealish in 2025. Grealish's claim that Reck 'has not contributed very much' is directly contradicted by Reck's own 2023 submission PDF. This is the clearest evidence that the TOTY process is being used to protect institutional interests rather than honor community service.",
     "VP Report + 2023 TOTY PDF (physical document analyzed) + two-cycle targeting pattern"),

    # ── UNAUTHORIZED SPENDING — NO APPROVED BUDGET ────────────────────────
    ("F-14", "Jul 1 – Nov 7, 2025", "Unauthorized Spending", "Legal + Governance",
     "HIGH", "Jeanne Grealish",
     "Standard nonprofit governance — expenditures require an approved budget; PMTNM Handbook: budget must be approved by Executive Board at a board meeting; NM Nonprofit Corporation Act — fiduciary duty",
     "PMTNM's fiscal year 2025-2026 began July 1, 2025. Grealish's own Oct 7 email confirms the CPA report was not expected until 'around November 1st' because the CPA 'was out of the country all summer.' The proposed budget was not sent to the budget committee until Oct 29, 2025 — four months into the fiscal year. A formal budget committee meeting never took place (replaced by emails, per Grealish's direction). The budget was presented at the Nov 7 board meeting, but no formal vote was documented in the audio recording (consistent with M-02).\n\nDuring the Jul 1 – Nov 7 period, Grealish continued spending organizational funds — including $4,600 in storage costs (F-10), her $1,750 annual honorarium (F-07), and operational expenses — without any board-approved budget authorizing those expenditures. Grealish justified this by saying 'Conference expenses and all expenses in the 2025-2026 fiscal year are covered in the current budget' (Oct 7 email) — but no 2025-2026 budget existed yet. The only operative document was the prior year's budget, which was itself adopted under disputed circumstances.",
     "Oct 7, 2025 Grealish email (thread ID: 199c0d60237afa6c): 'The budget for 2026-2027 will be prepared once the accountant completes the year end report...That report will be available around November 1st'; 'No harm done since any new budget will not become affective [sic] July 1, 2026'; Oct 29, 2025: CPA report and draft budget distributed — 4 months after fiscal year start; Nov 7, 2025: budget 'voted on' but no formal vote in audio recording (see M-02); DocsFromJeanne financial records showing ongoing expenditures during Jul–Nov 2025",
     "Four months of organizational spending with no board-approved budget. Grealish dismissed the problem by claiming 'no harm done' — while simultaneously blocking all committee oversight. Combined with F-05 (sole financial control) and F-06 (personal benefit from storage), the 4-month gap represents completely unaccountable spending by a single officer.",
     "Email-documented timeline — fiscal year gap confirmed"),

    # ── AUDITING COMMITTEE DISMANTLED ─────────────────────────────────────
    ("A-01", "Ongoing (confirmed Oct 30, 2025)", "Auditing Committee Abandoned", "Legal + Governance",
     "HIGH", "Jeanne Grealish",
     "PMTNM Handbook 2004 Section VII-D (Auditing Committee): 'Makes a complete audit of the books prior to (if possible) or during the State Conference. All figures in the ledgers, journals, check book, receipts, and vouchers should be checked. Makes a report at the State Conference Business Meeting.'",
     "In her private letter to attorney Kathy Black (Oct 30, 2025), Grealish stated: 'PMTNM does not have a separate Finance Committee. It has not had an Auditing Comm. since hiring a CPA to double check all activity and file all required reports with IRS and the State of NM.' This is a written admission that a standing committee required by the PMTNM Handbook has been unilaterally eliminated.\n\nThe distinction is legally critical: a CPA preparing tax returns and an annual fiscal report is NOT an audit. The handbook Auditing Committee is an internal oversight function — committee members verify the ledgers, journals, checkbook, receipts, and vouchers against each other and report at the State Conference. That independent check of the books has not existed for years. Kathy Black confirmed: 'The function of an audit committee should be to retain a CPA for a formal audit, which this organization does not do.'\n\nThe elimination of the Auditing Committee means Grealish's financial records — held exclusively at her home with no online access, no other signatories, and no independent verification — have gone unchecked by any committee member. The CPA she hired (Mary Scofield) prepares reports based solely on materials Grealish provides her.",
     "Grealish private letter to Kathy Black, Oct 30, 2025 (forwarded to Chip by Kathy Black, Gmail thread ID: 19a3309ed58127ac): 'It has not had an Auditing Comm. since hiring a CPA to double check all activity and file all required reports with IRS and the State of NM'; Kathy Black Oct 30, 2025: 'The function of an audit committee should be to retain a CPA for a formal audit, which this organization does not do'; 2004 Handbook Section VII-D: verbatim Auditing Committee duties requiring complete audit at State Conference; CPA Mary Scofield — prepares tax returns only, not independent audit",
     "Standing committee required by the handbook has been eliminated without board vote. The internal financial oversight function that would have caught irregularities in storage expenses (F-10), personal property co-mingling (F-06), and sole-signatory arrangements (F-05) has not operated for years. Grealish's justification — hiring a CPA — is legally incorrect as preparation of tax returns is not an independent audit of the books.",
     "Grealish's own written admission — attorney-forwarded"),

    # ── MINUTES IN STORAGE / RECORDS ACCESS ───────────────────────────────
    ("M-06", "Oct 30, 2025 (confirmed in writing)", "Records in Storage — Minutes Inaccessible", "Legal + Governance",
     "HIGH", "Jeanne Grealish",
     "NM Nonprofit Corporation Act NMSA §53-8-26 — records must be available for inspection; PMTNM Handbook (2004): 'The records should be open to any member at all times' (EST duties section, re: Auditing Committee); Standard nonprofit governance — minutes are the authoritative record of corporate actions",
     "In her private letter to Kathy Black (Oct 30, 2025), Grealish confirmed: 'Past records and tax documents (both federal and state) are kept securely in storage along with Minutes from Board and General Meetings.' The board minutes — the organization's official legal record of all votes, decisions, and actions — are stored in the same storage unit where Grealish also stores personal furniture (F-06). Current financial records (CPA reports, checkbook, bank statements, deposit records, membership records) are kept at her personal home office.\n\nThis means: (1) Past meeting minutes cannot be accessed without retrieving them from the storage unit. (2) Current financial records cannot be accessed without visiting Grealish's home. (3) 'PMTNM does not do online banking' — no electronic access exists. (4) When Kathy Black formally demanded records by Oct 31, 2025, Grealish said it was 'not physically possible' to comply. This is a textbook records-access failure under NMSA §53-8-26.",
     "Grealish private letter to Kathy Black, Oct 30, 2025 (Gmail thread ID: 19a3309ed58127ac): 'Past records and tax documents (both federal and state) are kept securely in storage along with Minutes from Board and General Meetings'; 'Current files (CPA reports, checkbook, bank statements for all seven accounts, deposit records, and membership records) are kept in my home office at 1226 Morningside Dr. NE'; 'PMTNM does not do online banking'; 'It is not physically possible for me to bring the requested items to your office on Friday, October 31, 2025'; Kathy Black letter to Grealish (Oct 29, 2025) citing NMSA §53-8-27: formal demand for inspection of records",
     "Board meeting minutes — the organization's legal record of its decisions — are locked in a storage unit controlled solely by Grealish. When an attorney formally demanded records under NM statute, Grealish said it was physically impossible to produce them in a business day. Combined with F-05 (sole control), this means no board member, member, or attorney can independently verify any of PMTNM's financial history without Grealish's physical cooperation.",
     "Written admission in attorney-forwarded letter — direct statutory violation"),

    # ── BUDGET COMMITTEE — SPRING TIMING VIOLATION ────────────────────────
    ("F-15", "Spring 2025 (failure to act)", "Budget Timing Violation", "Governance",
     "MEDIUM", "Jeanne Grealish; Sharon Kunitz",
     "PMTNM Handbook 2004 Section VII-E (Budget Committee): 'Meets before the Spring Board Meeting each year to prepare a budget for the coming year and presents it to the Executive Board for approval.'",
     "The PMTNM Handbook requires the Budget Committee to meet BEFORE THE SPRING BOARD MEETING to prepare the budget for the coming fiscal year. PMTNM's fiscal year runs July 1 – June 30. For FY 2025-2026, the budget should have been prepared in spring 2025 (before a spring board meeting), approved by the Executive Board at that meeting, then recommended to the membership at the State Conference.\n\nInstead: No spring board meeting was held in 2025. No budget committee convened in spring 2025. The CPA was 'out of the country all summer.' The budget was not even drafted until October 29, 2025 — four months into the fiscal year — and was based on figures that only arrived October 26. The organization operated without any approved budget from July 1 through at least November 7, 2025.\n\nThis is not a one-time delay. The pattern of fall-only board meetings and no spring budget process appears to be the standard practice under Grealish's administration.",
     "Oct 7, 2025 Grealish email (thread ID: 199c0d60237afa6c): 'The budget for 2026-2027 will be prepared once the accountant completes the year end report... That report will be available around November 1st'; 2004 Handbook Section VII-E: 'Meets before the Spring Board Meeting each year'; No spring 2025 board meeting documented; Budget draft sent Oct 29, 2025 — 4 months into FY 2025-2026; Grealish Oct 7: 'No harm done since any new budget will not become affective [sic] July 1, 2026' [note: this confusingly refers to next year's budget, showing Grealish herself was confused about the timeline]",
     "The handbook-required spring budget process has been systematically bypassed. No spring board meeting, no spring budget committee, no pre-fiscal-year budget approval. Four months of unauthorized spending under no approved budget. The person responsible for calling the committee is also the person whose spending would have been scrutinized by that committee.",
     "Documentary — handbook vs. actual calendar"),

    # ── MINUTES FALSIFICATION — BUDGET 'APPROVAL' THAT NEVER OCCURRED ─────────
    ("M-07", "November 7, 2025", "Minutes Falsification", "Legal + Ethics",
     "HIGH", "Jeanne Grealish (author/distributor); Sharon Kunitz (presiding officer)",
     "Fiduciary duty to maintain accurate organizational records; Nonprofit corporate record standards; State law on corporate minute accuracy; Robert's Rules of Order — accurate record of motions and votes",
     "The November 7, 2025 board meeting minutes (distributed Mar 9, 2026, four months late) state: 'Secretary Treasurer's Report: The current report with financial summary was approved as presented.'\n\nThe audio recording at timestamp 25:52 shows that Grealish mentioned the budget was 'prepared for 26/27' as a passing remark tacked onto the end of the designated accounts motion — no document was presented, no figures discussed, no motion made, no second, no vote. The only thing formally voted on in the Secretary-Treasurer's report section was the Mary Scofield CPA designated accounts motion, which the minutes do record separately.\n\nThe 'approved as presented' language for the financial summary and budget is FABRICATED and appears nowhere in the audio recording. This fictitious approval retroactively authorizes Grealish's $1,750 annual honorarium for another fiscal year without a real vote. Chip had raised written questions about the honorarium on October 31, 2025 ('Has your honorarium increased through the years as our membership has decreased? What are the measures for your honorarium?') — questions that were never answered. The false minutes entry serves to lock in that compensation without board action.",
     "Audio Recording Nov 7, 2025, timestamp 25:52 — Grealish's remark without any motion, second, or vote; Nov 7 Minutes (distributed Mar 9, 2026) — falsified 'approved as presented' language; Oct 31, 2025 email — Chip's written questions about honorarium (thread ID: 19a3309ed58127ac); Audio recording ~46 minutes total — full context of Secretary-Treasurer's report section (no other votes on budget/financial summary)",
     "Minutes state a budget and financial summary were 'approved as presented,' but the audio recording proves no such motion or vote occurred. Minutes are a binding legal record of board action. A false entry that retroactively authorizes officer compensation constitutes falsification of corporate records — particularly when the fabrication appears 4 months after the meeting, after litigation has commenced.",
     "Audio-verified falsification — fabricated approval language contradicted by full recording"),

    # ── NM SECRETARY OF STATE — ANNUAL REPORT 6+ YEARS OVERDUE ──────────
    ("F-16", "Due Nov 15, 2019 — still overdue as of Oct 2025", "Regulatory Failure + False Statement", "Legal + Ethics",
     "CRITICAL", "Jeanne Grealish",
     "NM Nonprofit Corporation Act — annual report filing requirement with NM Secretary of State; Fiduciary duty of accurate regulatory compliance; Prohibition on false statements in official organizational records",
     "On October 23, 2025, attorney Kathy Black reviewed PMTNM's NM Secretary of State business records and found: 'PMTNM's annual report is way overdue, with the due date listed as November 15, 2019.' The NM Secretary of State annual report has not been filed in over SIX YEARS.\n\nThis directly contradicts a statement in PMTNM's own internal financial documents (Budget Documents, drafts approved by Board): 'All federal, state and local required documents and lists have been prepared and filed.' If this statement appeared in a report adopted by the board, it constitutes a false certification in an official organizational record — either Grealish failed to file and knew it, or she represented compliance to the board falsely.\n\nNote: Kathy Black noted this same lapse on Oct 30, 2025 when she said 'At the very least, I am going to point out that the State of New Mexico annual report has not been filed.' That was her second reference to this same failure — she had already identified it on Oct 23.\n\nAlso relevant: the NM AG Charitable Organization Registration (F-12) listed the organization as filing a 990-N and reporting 'solicited funds in NM: No.' The Secretary of State lapse compounds the regulatory picture: PMTNM has not maintained compliance with the most basic state corporate filing requirement since 2019, while Grealish continues to execute contracts (F-13: VIP Staffing, Oct 2025) and spend organizational funds as if the organization is in good standing.",
     "Kathy Black email to Chip Oct 23, 2025: 'according to New Mexico Secretary of State business records, PMTNM's annual report is way overdue, with the due date listed as November 15, 2019'; Kathy Black email Oct 30, 2025: 'At the very least, I am going to point out that the State of New Mexico annual report has not been filed'; PMTNM internal document ('Budget Documents, drafts approved by Board (1).pdf' in Drive — PMTNM folder): 'All federal, state and local required documents and lists have been prepared and filed' — DIRECT CONTRADICTION of Secretary of State records; NM Secretary of State public business records (verifiable)",
     "PMTNM has been operating for 6+ years without filing its required NM Secretary of State annual report. Its own internal financial documents affirmatively claim all required filings are current — a false statement in an official record. This is simultaneously a regulatory failure, a false representation to the board, and evidence that Grealish's self-reported compliance ('The Executive Secretary-Treasurer shall be responsible for... filing of all returns with IRS, NM Taxation and Revenue Department, the NM State Corporation Commission, and any other government entity') has not been fulfilled.",
     "Attorney-identified + contradicted by PMTNM's own internal document"),

    # ── NM AG REGISTRATION — TWO FILING YEARS — PATTERN CONFIRMED ────────
    ("F-12", "Tax Year 2018 (filed ~2019); Tax Year 2023 (filed Dec 18, 2024); Tax Year 2024 (filed Dec 12, 2025)", "Regulatory / Sole Control", "Legal + Governance",
     "CRITICAL", "Jeanne Grealish",
     "NM Attorney General Charitable Organization Registration requirements; IRS NTEE classification accuracy; NM Solicitation Disclosure Act; NM Nonprofit Corporation Act — officer record accuracy",
     "THREE YEARS OF NM AG FILINGS REVIEWED — IDENTICAL SOLE-CONTROL PATTERN ACROSS DECADE:\n\nTAX YEAR 2018 (filed ~2019):\n(1) Grealish sole responsible party for ALL FIVE financial functions. (2) PMTNM registered address = Grealish's personal home. (3) NTEE misclassification: A6B (Singing & Choral Groups) — incorrect for a music TEACHERS professional association. (4) Revenue $18,817 vs. expenses $22,649 — operating deficit. (5) 'Solicited funds in NM: No' — appears false; PMTNM collects dues, conference fees, foundation contributions.\n\nTAX YEAR 2023 (FY 7/1/2023–6/30/2024, submitted Dec 18, 2024, status 'Registration Submitted' — not fully approved):\n(1) Grealish STILL sole responsible party for ALL FIVE functions: check signer, fund raising, fund distribution, financial records custody, and fund custody — no other officer listed for any function. (2) PMTNM address STILL Grealish's personal home: 1226 Morningside Dr NE, Albuquerque NM 87110. All officers listed at same personal address. (3) Officers listed: Larry Blind (President, $0), Jeanne Grealish (EST, $1,750), Cheryl Pachak-Brooks (Vice President, $0). CHIP MILLER NOT LISTED AS VP — but this is technically accurate: Chip became VP in August 2024, which falls in Tax Year 2024 (FY 7/1/2024–6/30/2025), not Tax Year 2023. Cheryl Pachak-Brooks WAS VP during FY 2023-2024. The filing is period-accurate. (4) 990-N filer. Revenue $15,819; Expenses $15,589; Net Assets end of year: $62,500; Contributions: $1,854. (5) Accountant: Mary Scofield, self-employed (not a CPA firm), 1308 Alcazar NE — all financial reporting routed through a sole self-employed accountant with no institutional independence. (6) 'Solicited funds in NM: No' — STILL false; Solicitation Methods listed as 'Special Events, Personal Contact' — direct internal contradiction within the same filing. (7) NTEE still includes A6B.\n\nTAX YEAR 2024 (FY 7/1/2024–6/30/2025, filed Dec 12, 2025 — after lawsuit filed Jan 27, 2026):\n(1) All five financial functions REMAIN solely Grealish. (2) Net assets declined from $62,500 to $60,555 (continuing decade-long decline from $81,975 in 2014 — total erosion of $21,420 or 26%). (3) Total gross revenue $7,114 (down 70% from $23,963 in 2014). (4) Total contributions $0.00 despite 'Special Events' and 'Personal Contact' listed as solicitation methods. (5) 'Solicited funds in NM: No' answered false for the eleventh consecutive year — systematic false statement across the entire regulatory history. (6) Chip Miller (Vice President) listed on filing. (7) This filing was submitted after the lawsuit was filed, suggesting a document created under legal awareness.",
     "NM AG Registration Tax Year 2018 (physical PDF: 20184721935751263.pdf); NM AG Registration Tax Year 2023 (physical PDF: 20234722435369467 (2).pdf); NM AG Registration Tax Year 2024 (filed Dec 12, 2025), FEIN 85-0284938; VP appointment email from Sharon Kunitz, Aug 14, 2024 — establishes Chip's VP status; IRS BMF confirming 990-N filer; Internal financial records (DocsFromJeanne) and Kathy Black Initial Impressions confirming same sole-control pattern 2024-2025",
     "Three independent NM AG filings across a decade show IDENTICAL sole-control structure. The 'solicited funds in NM: No' false statement appears in all three filings despite documented solicitation activity — a systematic misrepresentation maintained throughout the regulatory history. The Tax Year 2024 filing (submitted after the lawsuit was filed) documents that Chip Miller was the VP during that year, yet all financial control remained with Grealish. Net assets and revenue continue their decade-long decline while officer compensation (Grealish honorarium) increased from 6.3% to 24.6% of revenue — compensation rising as the organization shrinks.",
     "THREE physical documents reviewed — entire decade pattern confirmed at NM AG level"),
    # ── VIP STAFFING CONTRACT — UNAUTHORIZED COMMERCIAL AGREEMENT ─────────
    ("F-13", "Oct 23, 2025", "Unauthorized Contract", "Legal + Governance",
     "HIGH", "Jeanne Grealish",
     "Nonprofit governance — contracts above a de minimis threshold require board authorization; PMTNM Handbook on financial authority; Fiduciary duty",
     "On October 23, 2025, PMTNM entered into a Master Terms & Conditions agreement with VIP Staffing — a commercial staffing agency. Three versions of this contract are preserved in Google Drive (draft, executed, and a third copy), all dated Oct 23, 2025. PMTNM is identified as the 'Customer.' For a nonprofit that (a) filed 990-N (revenue under $50K threshold), (b) had FY2018 revenue of only $18,817, (c) claims in the NM AG filing it does not solicit funds, and (d) operates from Grealish's personal home — entering a commercial staffing contract is a significant financial and legal commitment. No board vote authorizing this contract appears in any board minutes. The Nov 7, 2025 board meeting (the only documented board meeting during this period) contains no reference to a staffing contract. Grealish, as sole financial officer, appears to have executed this agreement unilaterally.",
     "Google Drive: '2025-10-23T16-25-09.000Z - 2025 VIP Staffing Master Terms Conditions- Professional Music Teachers of New Mexico.pdf' (Drive ID: 1OGc7fLQguB3-ZDiqXpuUxF83sK2Rcy-L); '2025-10-23T16-49-03.000Z - Executed 2025 VIP Staffing Master Terms Conditions- Professional Music Teachers of New Mexico.pdf' (Drive ID: 1nQAsasqmpAAk9d3-ZZJum3j4a70G92kZ); Nov 7, 2025 Board Meeting minutes and audio — no reference to any staffing contract; No board resolution authorizing staffing agreement found",
     "A commercial staffing contract executed without board authorization — by the sole financial officer of a nonprofit operating at a deficit. The contract requires PMTNM to carry Comprehensive General Liability Insurance and Employment Practices Liability Insurance. Whether PMTNM actually maintains this insurance is unknown. The agreement also contains a one-year non-solicitation clause that would bind PMTNM as an organization. This is precisely the kind of significant financial obligation that requires board authorization under standard nonprofit governance.",
     "Physical contracts reviewed in Drive — no board authorization documented"),

    # ── OFFICER COMPENSATION WITHOUT BOARD AUTHORIZATION ──────────────────────
    ("F-17", "Ongoing (documented 2014–2025); FY2024–25 Authorization Denied", "Financial / Officer Compensation", "Legal + Governance",
     "CRITICAL", "Jeanne Grealish (recipient); Sharon Kunitz (presiding officer, failed oversight); Larry Blind (Nominating Chair, failed oversight)",
     "Nonprofit governance — officer compensation requires board approval; PMTNM Handbook duties for EST and President; Fiduciary duty to manage organizational resources; NM corporate law — board's duty to authorize material expenses",
     "Grealish (Executive Secretary-Treasurer) has received annual compensation ('honorarium') of at least $1,500 (Tax Year 2014) rising to $1,750 (Tax Year 2024), disclosed in NM AG filings but never appearing in IRS records because PMTNM files 990-N. The sole financial controller of all five functions (check signing, fundraising, fund distribution, financial records, custody of funds) is the same person receiving this compensation. No independent oversight exists.\n\nIn fiscal year 2024–25 (the year Chip served as VP), the budget containing this compensation was never formally voted on by the board:\n— The 'Budget Committee REMINDER' process (Oct 31–Nov 6, 2025) was entirely email-based, not a formal meeting\n— Larry Blind approved it by reply-all email on Nov 2 ('The budget looks ok to me')\n— The November 7 board meeting contained no motion, second, or vote on the budget — audio confirmed at timestamp 25:52 (passing remark, not a motion)\n— The minutes' 'approved as presented' language is falsified (see M-07)\n\nSince the budget was never formally approved by the board in FY2024–25, Grealish's $1,750 honorarium for that year was paid without authorization.\n\nSALARY VS. REVENUE TRAJECTORY (from NM AG filings):\n— Tax Year 2014: $1,500 salary / $23,963 revenue = 6.3%\n— Tax Year 2024: $1,750 salary / $7,114 revenue = 24.6%\n\nAs membership and revenue collapsed 70% over a decade, Grealish's share of organizational revenue nearly quadrupled. The Tax Year 2024 NM AG filing (submitted December 12, 2025 — after the lawsuit was filed) shows $1,750 compensation resting on a board approval the audio recording and minutes contradiction prove never happened.",
     "NM AG Registration filings Tax Year 2014–2024 (pattern documented); Oct 31–Nov 6, 2025 'Budget Committee REMINDER' email thread (Gmail ID: 19a3309ed58127ac); Nov 7, 2025 Audio Recording, timestamp 25:52 — no motion or vote on budget; Nov 7, 2025 Minutes (distributed Mar 9, 2026): falsified 'approved as presented' language (see M-07); NM AG Tax Year 2024 filing (submitted Dec 12, 2025) showing $1,750 compensation; Revenue data from IRS EO BMF and NM AG filings 2014–2024",
     "A manager received continuous compensation without formal board authorization in the final fiscal year reviewed. The email-based budget process bypassed handbook requirements (F-02). The Nov 7 board meeting produced no motion to approve the budget — only a fabricated minutes entry (M-07). The compensation increased from 6.3% to 24.6% of declining revenue, while the sole financial officer remained unilaterally in control. This is foundational financial governance failure.",
     "Audio-verified (no motion), Minutes-contradicted (fabricated approval), NM AG-documented (11-year pattern)"),

    # ── VOTE OF NO CONFIDENCE — MOTION MADE, PROCEDURALLY DEFEATED ─────────
    ("G-02", "Nov 7, 2025 (pre-meeting emails + meeting)", "Anonymous Voting Suppressed + No Confidence Motion Defeated", "Governance + Ethics",
     "CRITICAL", "Larry Blind (procedural suppression); Sharon Kunitz (adjournment); Jeanne Grealish (subject of motion, present during vote call)",
     "Robert's Rules of Order Newly Revised 12th ed. §§45:5–7, 46:32–34 (ballot voting must preserve anonymity); Robert's Rules §4 (main motions); Nonprofit governance — board members' duty to consider officer accountability motions",
     "TWO-PART VIOLATION — PRE-MEETING SUPPRESSION AND IN-MEETING PROCEDURAL DEFEAT:\n\nPART 1: ANONYMOUS VOTING SYSTEM PREPARED AND IGNORED (Nov 7, morning, Gmail thread: 19a5ea6717e3582c)\nAt 8:54 AM, Chip proposed two third-party anonymous voting platforms (eballot.com and electionchamp.com), citing RROO §§45:5–7, 46:32–34, and offered to pre-load all board member emails. At 10:00 AM, Larry replied ONLY to Sharon\'s original email (ignoring Chip\'s anonymous voting proposal) with: \'There may be a lot of talking, but getting someone to the point of a motion can be challenging\' — telegraphing his plan to prevent the motion procedurally, hours before the meeting. At 10:11 AM, Larry counter-proposed text votes to a single phone number (\'designate one person to receive the texts, count the results, and then delete the texts\') — not anonymous by RROO standards because the designated receiver sees all votes. Chip at 10:27 AM explained why this fails the ballot vote standard. At 10:55 AM, Chip confirmed: \'ElectionChamp is set up. It is straightforward.\' He sent a test election to all 15 board member emails. The system was live and ready.\n\nPART 2: MOTION DEFEATED BY PUBLIC CALL IN HOSTILE ENVIRONMENT (Nov 7 meeting, ~46:34–49:13)\nDespite ElectionChamp being set up and waiting, Larry ran the motion of no confidence as a public floor call — asking all attendees to visibly raise their hands or speak affirmatively to second the motion. This was done:\n- With Grealish physically present as the subject of the motion\n- With Larry presiding despite having sent Chip a hostile email the night before (L-04)\n- With all in-person attendees fully visible to each other\n- Without any reference to the anonymous voting system Chip had prepared\nNo one seconded. Motion died. Sharon adjourned immediately.\n\nLarry\'s 10:00 AM comment (\'getting someone to the point of a motion can be challenging\') was written before the meeting as a preview of his strategy — not a neutral procedural observation. The anonymous voting system was the correct remedy under RROO and was available. Its non-use ensured the vote occurred under maximum social pressure.\n\nTRIGGER CONFIRMED: The Nov 6 pre-meeting coordination (Larry's 10:00 AM anonymous voting rejection, ElectionChamp setup by 10:55 AM, Cheryl briefed by evening) was preceded by Grealish forwarding Chip's retirement request to MTNA national office (Brian Shepard) and her attorney (Karen Kilgore) on Nov 5, 2025 at 7:03 PM. The sequence shows: Chip asks Grealish to retire in Nov 5 meeting → Grealish escalates to MTNA and attorney same evening → coordinated pre-meeting suppression activity begins Nov 6 morning. The anonymous voting suppression was not spontaneous — it followed a deliberate escalation the prior evening.\n\n2024 PRECEDENT — PERSONAL CELL PHONE VOTING (Gmail thread 1928ce26b2fe3d9c, approx. Oct 14, 2024):\nIn 2024, the year before ElectionChamp was suppressed, the mechanism for online \'anonymous\' voting was: board members should text their votes to Larry Blind\'s personal cell phone (505) 980-3941. The meeting was held at NM School of Music — Larry\'s own commercial school.\n\nWHY THIS IS NOT ANONYMOUS: Larry, as the sole recipient, could see every vote cast — identical in structure to his Nov 7, 2025 counter-proposal (\'designate one person to receive the texts\') that Chip rejected as non-anonymous under RROO §§45:5–7, 46:32–34. A vote-receiver who sees all votes before counting cannot provide the anonymity Robert\'s Rules requires for ballot votes.\n\nPATTERN ACROSS TWO CONSECUTIVE ELECTIONS: True anonymous voting was never available to online board members under Larry\'s administration. 2024: votes texted to Larry\'s personal cell at his own school. 2025: Chip\'s RROO-compliant ElectionChamp system (15 board members loaded, ready by 10:55 AM) was rejected by Larry that morning; vote run publicly instead with Grealish in the room. Two election cycles, zero genuine anonymous voting for online members. The 2025 suppression was not an oversight — it was a continuation of established practice.",
     "Gmail thread 19a5ea6717e3582c — full pre-meeting exchange: Sharon 7:07 AM (time limits); Chip 8:54 AM (anonymous voting proposal, RROO citation, two platforms offered); Larry 10:00 AM (dismissive, to Sharon only: \'getting someone to the point of a motion can be challenging\'); Larry 10:11 AM (text-to-one-phone counter-proposal — not anonymous); Chip 10:27 AM (explains why Larry\'s method fails RROO ballot standards); Chip 10:55 AM (\'ElectionChamp is set up\'  — 15 board member emails loaded, test sent); Nov 7 Audio Recording 46:34–49:13: motion made, public hand-raise call, no second, motion dies; Sharon 49:13: adjourns; Gmail thread 1928ce26b2fe3d9c — \'PMTNM Board and General Membership Meetings\' (approx. Oct 14, 2024): Larry announces online members should text votes to personal cell (505) 980-3941; meeting held at NM School of Music (Larry\'s school) — establishes 2-year pattern of non-anonymous online voting",
     "A proper anonymous voting system was live and ready. Larry rejected it that morning in favor of a non-anonymous alternative, then ran the vote publicly with the subject of the motion in the room. Members who might have seconded the motion anonymously were instead exposed to full social and professional visibility. The motion\'s failure cannot be separated from the voting environment Larry created and the anonymous option he suppressed.",
     "Email-verified pre-meeting + Audio-verified in-meeting — anonymous voting deliberately bypassed — Trigger: Nov 5, 7:03 PM forward to MTNA/attorney preceded Nov 6 coordination"),

    # ── LARRY BLIND — UNDISCLOSED TRIPLE ROLE ─────────────────────────────────
    ("L-07", "Nov 7, 2025 (and ongoing 2024-2025)", "Conflict of Interest / Structural Governance", "Ethics + Governance + Procedural",
     "CRITICAL", "Larry Blind (three simultaneous conflicting roles); Sharon Kunitz (failed to manage or disclose)",
     "PMTNM Handbook 2004 Section VII-Q (Nominating Committee chaired by Immediate Past President — controls VP succession); Section VI-C (President duty: 'keeps a close liaison with all officers' and should allocate duties to VP — not fulfilled while receiving exclusive advisory counsel from Past President); Robert's Rules of Order — Parliamentarian must be impartial and free of conflicts; Standard nonprofit governance — board members holding multiple conflicting roles must disclose them",
     "At the November 7, 2025 board meeting and in the weeks leading up to it, Larry Blind simultaneously held THREE distinct roles — none of which were disclosed as conflicting and none of which had any handbook-defined boundaries or recusal process:\n\n(1) PAST PRESIDENT / INFORMAL ADVISOR TO SHARON: The 2004 PMTNM Handbook Section VI defines duties for State President, Vice President, EST, and District VPs. The Past President is NOT defined in Section VI. The advisory relationship is conventional — meaning no handbook language defines its scope, limits, or conflict-of-interest requirements. This gave Larry an unconstrained, informal advisory channel to Sharon with no procedural safeguards.\n\n(2) NOMINATING COMMITTEE CHAIR (per Section VII-Q convention): The Past President chairs the Nominating Committee — the body that recommends candidates for future elected officer positions, including VP. Larry controls VP succession while simultaneously acting as the current VP's adversary.\n\n(3) PARLIAMENTARIAN AT NOV 7 MEETING: Larry served as the sole parliamentary authority at the meeting where he had already been advising Sharon against Chip's agenda proposals that morning.\n\nTHE ADVISORY CHANNEL WAS ACTIVELY USED AGAINST THE VP: At 10:00 AM on Nov 7, Larry replied to Sharon's original email — deliberately ignoring Chip's anonymous voting proposal — with: 'There may be a lot of talking, but getting someone to the point of a motion can be challenging.' This was not sent to Chip. It was an advisory communication to the President about how to manage the meeting's parliamentary agenda hours before it began. The anonymous voting system Chip had set up by 10:55 AM was never used. The no-confidence motion was instead called by public hand-raise, in a room with Grealish physically present, under Larry's parliamentary authority — and died without a second.\n\nSHARON'S OWN HANDBOOK DUTY: Section VI-C requires the President to 'keep a close liaison with ALL officers' and allocate duties to the VP. Sharon demonstrably did not do this while receiving Larry's informal counsel. The advisory channel substituted for the VP-President relationship the handbook was designed to foster.\n\nNO CONFLICT DISCLOSURE, NO RECUSAL, NO PROCEDURE: There is no PMTNM conflict-of-interest policy in the handbook. No disclosure of Larry's three simultaneous roles was made at the Nov 7 meeting. No recusal procedure existed. Larry ruled on parliamentary questions at the same meeting where he had been advising the President against the VP and where he controlled the VP successor nomination process.",
     "2004 PMTNM Handbook Section VII-Q (Nominating Committee: Past President chairs by convention — not formally defined in Section VI); 2004 Handbook Section VI-C (President duties: 'Keeps a close liaison with all officers' — not extended to VP during Larry's advisory relationship); Nov 7, 2025 10:00 AM email from Larry — reply to Sharon only, not Chip (Gmail thread 19a5ea6717e3582c): 'There may be a lot of talking, but getting someone to the point of a motion can be challenging' — advisory communication bypassing VP; Nov 7 Audio Recording: Larry as Parliamentarian ruling on procedural questions; Nov 7 Minutes: Larry named as Parliamentarian; Nov 5-6 hostile email from Larry to Chip (L-04): 'offensive, disrespectful, and does not deserve a response' — demonstrates pre-existing personal animus that preceded his impartial parliamentary role; No conflict-of-interest policy found in 2004 PMTNM Handbook",
     "Three simultaneous conflicting roles — advisor, succession controller, parliamentary authority — held without disclosure or recusal. The informal advisory channel had no handbook boundaries. Larry's 10:00 AM email proves the advisory channel was used to coordinate meeting strategy against the VP hours before the meeting. Parliamentary rulings at Nov 7 were made by a person who had been advising the President in that direction since at least 10 AM. The Nominating Committee process is further compromised by Larry's documented financial conflicts (L-01 through L-06) and his use of organizational roles for personal and commercial benefit.",
     "Handbook-verified + Email evidence (10:00 AM advisory communication) + Audio record — three undisclosed simultaneous roles"),

    # ── LARRY BLIND — UNILATERAL CONFERENCE PROGRAMMING DECISION ──────────────
    ("G-05", "Oct 4, 2023 — Oct 9, 2023", "TOTY Process Monopoly — Employee Administers Vote on Boss's Nomination", "Governance + Procedural",
     "CRITICAL", "Lawrence Blind",
     "Nonprofit governance — conflicts of interest must be managed through recusal or independent administration; Robert's Rules of Order — procedural fairness; NM nonprofit law — fiduciary duty",
     "Larry Blind, an NMSM employee, held three simultaneous operational roles in the 2023 PMTNM Teacher of the Year process:\n\n(1) PMTNM PRESIDENT: Convened the voting process and set procedural rules (Oct 4, 2023 email: 'I decided to extend the nomination period')\n\n(2) WEBSITE ADMINISTRATOR/PLATFORM CONTROLLER: Managed the voting platform (Google Form link: https://forms.gle/LqpzibHS4V3tJVWK9), set the deadline unilaterally ('You must register your vote by Monday, October 9'), and controlled communications to voters\n\n(3) VOTE RECEIVER/CONTACT PERSON: Was the sole contact for voting questions ('Text me at 505-980-3941 if you have any questions or trouble voting')\n\nSIMULTANEOUSLY, his BOSS (Tatiana Vetrinskaya, NMSM co-founder/owner) was the winning TOTY nominee. Larry could not be impartial about administering a process that would determine whether his employer received professional honors. He had maximum incentive to ensure she won — his employment depends on her goodwill.\n\nNO RECUSAL PROCEDURE existed. No independent administrator was appointed. No conflict of interest was disclosed. The employee simply controlled the entire process involving his boss.",
     "Oct 4, 2023 email from lblind@nmschoolofmusic.com (PMTNM President, acting as sole administrator): 'I decided to extend the nomination period'; 'Please read the attached nomination forms...click on the link below to place your vote'; 'You must register your vote by Monday, October 9'; 'Text me at 505-980-3941 if you have any questions'; NMSM organizational documentation confirming Tatiana as owner, Larry as employee; L-01 documentation of boss-as-nominee relationship",
     "A single individual — Larry Blind, NMSM employee — controlled every aspect of a voting process that determined whether his boss would receive organizational honors. No procedural safeguards existed: no recusal, no independent administration, no conflict disclosure. This is the definition of a rigged process. The monopolistic control combined with Larry's personal interest in the outcome (Tatiana is his employer) made fair administration impossible.",
     "Procedurally-verified — monopolistic control of conflict-of-interest vote documented"),

    ("G-03", "Oct 28, 2023 (approx.)", "Unilateral Conference Decision", "Governance + Conflict of Interest",
     "MEDIUM", "Lawrence Blind",
     "Standard nonprofit governance — significant organizational decisions require board approval and oversight; PMTNM Handbook (silent on conference programming authority, indicating governance gap); Fiduciary duty of loyalty — using organizational position and resources for personal/institutional benefit without disclosure",
     "Larry Blind, as PMTNM president, unilaterally decided to invite non-PMTNM students from his own school (NM School of Music) to participate in a PMTNM conference benefit concert scheduled for November 12, 2023. The decision appears to have been communicated directly to PMTNM member teachers in late October 2023. No board vote was recorded. No board meeting minutes document board discussion or approval of this programming decision. The handbook is silent on conference programming authority, leaving unclear whether this was within the president's delegated power or required board approval. Regardless, the decision involved a direct conflict of interest (his own school) that was neither disclosed nor managed through any recusal or conflict-of-interest procedure. The fact of NMSM students' participation in the 2023 PMTNM conference is independently corroborated in L-02 (ethics complaint filed Nov 14, 2023 and conference program documentation).",
     "Board meeting minutes from Oct/Nov 2023 — no documentation of board discussion, motion, or vote regarding conference programming or NMSM student participation; PMTNM 2004 Handbook — contains no section defining decision-making authority for conference programming or board approval requirements for conference content decisions; Independent corroboration: L-02 (Nov 14, 2023 ethics complaint to MTNA) and conference program/documentation from 2023 confirming NMSM student participation in PMTNM conference",
     "Unilateral decision-making by a president without board oversight. Direct self-dealing: using PMTNM's conference platform to promote his own school and student base. Absence of conflict-of-interest management or disclosure. Pattern-consistent with leadership's broader governance approach (see G-01: 'PMTNM does not function under any legal mandates'). Even if conference programming fell within the president's delegated authority, the absence of any disclosure mechanism for conflicts of interest represents governance failure.",
     "Documented via board minutes record (absence of decision) + independent corroboration via L-02 ethics complaint and conference documentation — Pattern consistent with unilateral leadership style"),

    # ── MARCUS YORK — UNMANAGED VOTING CONFLICT OF INTEREST ────────────────────
    ("L-08", "Nov 8, 2024; Nov 7, 2025", "Unmanaged Conflict of Interest / Proxy Voting", "Governance + Conflict of Interest",
     "HIGH", "Marcus York; Lawrence Blind (as supervisor/authority)",
     "Nonprofit governance — voting members with conflicts of interest must be subject to formal conflict-of-interest procedures (disclosure, recusal, abstention); PMTNM governance standards requiring impartial board member conduct; Proxy voting prohibition under standard nonprofit law; Fiduciary duty of loyalty",
     "Marcus York, an employee of NM School of Music (controlled by Lawrence Blind as owner/director), voted on critical financial matters at PMTNM board meetings while under employment relationship with Blind — with no formal conflict-of-interest procedure to manage or mitigate that relationship. Although PMTNM members likely knew York worked for Larry, the board took no formal action to address this conflict:\n\nNOVEMBER 8, 2024 BOARD MEETING (Blind was President):\n— Marcus York made a motion on the Treasurer's Financial Report (which was 'unanimously approved')\n— Marcus York seconded a motion on the Budget (which was 'declared in order')\n— York was physically in-person at the meeting held at NM School of Music (Blind's school)\n— No disclosure of NMSM employment relationship appears in the minutes\n— No conflict-of-interest statement or recusal was documented\n\nNOVEMBER 7, 2025 BOARD MEETING (Blind was no longer President but served as Parliamentarian):\n— Marcus York was listed as a member of the Nominating Committee, chaired by Lawrence Blind\n— The Nominating Committee controlled VP succession — a position Blind had just vacated and that was critical to the no-confidence motion against Grealish\n— York was not physically present at this Nov 7 meeting, but his continued appointment to the Blind-chaired Nominating Committee represents an ongoing undisclosed conflict\n\nTHE CONFLICT: York was voting (Nov 2024) on matters affecting the organization while being employed by someone with authority over him, while also serving on a committee chaired by that same person that controlled officer succession. NMSM's organizational authority structure gave Blind direct control over York's employment, creating the classic hallmark of proxy voting: a voting member whose interests are materially affected by someone else's authority.",
     "Board Minutes Nov 8, 2024: 'With a motion by Marcus York and a second by Sharon Kunitz the Treasurer's current financial report was unanimously approved'; 'Upon a motion by Janna Warren with a second by Marcus York [budget] was declared in order'; Meeting location: 'New Mexico School of Music, Albuquerque'; Board Minutes Nov 7, 2025: 'Nominating Committee Slate of Officers for 2026-2027: Chair Lawrence Blind thanked his committee members Astrid Groth and Marcus York'; Marcus York's NMSM email/employment confirmed in July 2024 communication identifying him as associated with 'the School of Music'",
     "A voting board member whose employment is controlled by someone else (Blind) voted on financial matters while serving on a succession-control committee chaired by that same person. Although people knew York worked for Larry, the board took no formal action to manage this conflict — no required disclosure, no recusal procedure, no abstention protocol. PMTNM has no written conflict-of-interest policy. This pattern extends Blind's demonstrated use of organizational positions to embed people aligned with his interests — similar to how he controlled the 2023 TOTY vote through his employee Tatiana. York's situation represents proxy-voting conditions: voting on matters affecting the organization while under employment relationship with an interested party, with no organizational safeguards.",
     "Board minutes verified — unmanaged employment conflict documented across voting and committee roles"),
]

# ── Write data rows ───────────────────────────────────────────────────────────
severity_colors = {
    "CRITICAL": (RED_BG, RED_TXT),
    "HIGH":     (GOLD_BG, GOLD_TXT),
    "MEDIUM":   (BLUE_BG, BLUE_TXT),
    "LOW":      (GRAY_BG, "404040"),
}

category_colors = {
    "Financial Control":    "1F3864",
    "Financial Self-Dealing": "C00000",
    "Financial Irregularity": "7F6000",
    "Conflict of Interest": "6B1F42",
    "Ethics Process":       "375623",
    "Minutes Falsification":"C00000",
    "Procedural Violation": "7F6000",
    "Governance Violation": "1F3864",
    "Broken Commitment":    "C00000",
    "Records Access":       "375623",
    "Records Distribution": "1F3864",
}

for i, row in enumerate(rows, 2):
    sev = row[4]
    bg, txt = severity_colors.get(sev, (WHITE, "000000"))
    for j, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.font = cell_font(color="000000" if j != 5 else txt, bold=(j == 5))
        c.fill = fill(bg if j == 5 else (GRAY_BG if i % 2 == 0 else WHITE))
        c.alignment = wrap()
        c.border = border

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [7, 14, 22, 20, 10, 25, 40, 55, 45, 40, 20]
for i, w in enumerate(col_widths, 1):
    set_col_width(ws1, i, w)

# Row heights
for i in range(2, len(rows) + 2):
    ws1.row_dimensions[i].height = 72

# ── Add table / auto-filter ───────────────────────────────────────────────────
last_row = len(rows) + 1
last_col = get_column_letter(len(headers))
tab = Table(displayName="Violations", ref=f"A1:{last_col}{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws1.add_table(tab)

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — SUMMARY BY PERSON
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("By Person")
ws2.freeze_panes = "A2"

persons = [
    ("Jeanne Grealish (EST)", RED_BG, RED_TXT, [
        ("F-01", "Financial records refusal — VP refused access to records (Mar 2025); attorney demand under statute refused (Oct 2025)"),
        ("F-02", "Budget committee bypassed — 'probably will be no need for an actual meeting' (Oct 7, 2025); handbook dismissed as 'grossly out of date'; attorney confirms non-compliance"),
        ("F-17", "Officer compensation without board authorization — $1,750 honorarium in FY2024–25 paid under falsified budget approval; compensation 24.6% of shrinking revenue (up from 6.3% a decade ago)"),
        ("F-14", "Unauthorized spending Jul–Nov 2025: 4+ months into fiscal year with no approved budget; continued storage, honorarium, and operating expenses without board authorization"),
        ("F-03", "CPA report 5 months late; budget adopted after fiscal year began"),
        ("F-04", "Financial report 'approved' in minutes — no motion/vote in recording"),
        ("F-05", "Sole control of all bank accounts, CPA contact, financial records"),
        ("F-06", "Personal furniture stored in PMTNM storage unit — admitted before attorney"),
        ("F-07", "EST honorarium $1,750/yr with no performance measures"),
        ("F-08", "Stated Teri Reck 'hadn't contributed' — potentially inaccurate/defamatory"),
        ("F-09", "Blocked membership booklet: '90 cents to mail' vs. 30%+ budget on storage"),
        ("F-10", "Storage unit: $4,600 actual vs. $3,000 budgeted — 53% overrun; even Larry questioned it"),
        ("F-11", "2025 TOTY: Three attempts to nullify Terri Reck nomination; directly contradicted by board members"),
        ("A-01", "Auditing Committee unilaterally eliminated — written admission to Kathy Black: 'It has not had an Auditing Comm. since hiring a CPA'"),
        ("M-06", "Board meeting minutes stored in storage unit (with personal furniture); attorney demand for records refused as 'not physically possible'; no online banking"),
        ("M-07", "Minutes falsification (Nov 7, 2025): budget 'approved as presented' — audio shows no motion, second, or vote; minutes fabricated to lock in unauthorized $1,750 honorarium"),
        ("F-15", "No spring board meeting or budget committee in spring 2025; budget prepared 4 months into fiscal year — handbook requires spring preparation"),
        ("F-16", "NM Secretary of State annual report overdue since Nov 15, 2019 — 6+ year lapse; internal PMTNM document claims 'all required documents have been filed' — direct contradiction"),

        ("F-12", "NM AG registration (2018, 2023, 2024): sole responsible party for all 5 financial functions all three years; PMTNM address = personal home all three years; 'no solicitation' false all three years; eleven-year pattern confirmed (2014-2024)"),
        ("F-13", "VIP Staffing contract (Oct 23, 2025): commercial staffing agreement executed without board authorization; no mention in Nov 7 minutes"),
        ("G-01", "ON RECORD Oct 7, 2025: 'PMTNM does not function under any legal mandates!'"),
        ("G-02", "Nov 7, 2025: vote of no confidence motion made formally; Larry controlled the second process; Sharon adjourned immediately after motion died; Grealish (subject of motion) had direct conflict of interest; escalation to MTNA/attorney Nov 5 preceded coordination"),
        ("M-01", "Minutes: '10-min limit' — Sharon's own email says 15 min"),
        ("M-02", "Minutes: financial report 'approved' — no vote in recording"),
        ("M-03", "Minutes: 'time called by Parliamentarian' — not in recording"),
        ("M-04", "Minutes: VP report 'not submitted' — contradicted by documentary evidence"),
        ("R-01", "Handbook withheld for 6 months; falsely claimed online availability"),
        ("R-02", "Nov 7 minutes distributed 4 months late — after petition filed"),
        ("R-03", "Spring 2025 minutes not distributed for 7 months"),
        ("R-04", "Official member list ('PMTNM. LIST AND EMAIL ADDRESSES') lists Sharon Kunitz as NCTM — contradicted by official MTNA data showing no NCTM designation; pattern of inaccurate record-keeping"),
    ]),
    ("Larry Blind (NMSM Employee / PMTNM President)", PURPLE_BG, PURPLE_TXT, [
        ("L-01", "TOTY conflict of interest: administering voting process for his BOSS Tatiana (NMSM owner) — can't be impartial when his employment depends on her goodwill"),
        ("G-05", "TOTY process monopoly: controlled nomination deadline, voting platform, and all communications while his boss was the nominee — no recusal, no independent administration"),
        ("L-02", "NMSM students at PMTNM conference — using nonprofit event to market his school"),
        ("L-03", "PMTNM website used to promote NM School of Music during his presidency"),
        ("L-04", "Served as Parliamentarian after sending hostile email to VP — compromised impartiality; Nov 7 pre-meeting email: 'getting someone to the point of a motion can be challenging' — telegraphed plan to prevent no-confidence motion; rejected Chip\'s RROO-compliant anonymous voting in favor of text-to-one-phone non-anonymous method; ran the motion vote publicly with Grealish in the room"),
        ("L-05", "TOTY 2023: solicited votes from NMSM email; winner is NMSM co-founder/employee; extended deadline unilaterally"),
        ("L-06", "NMSM biography: 'perform internationally' misleading; Carnegie Hall framing unsubstantiated; uncredentialed faculty"),
        ("E-01", "NMSM employee handbook invoked to block MTNA ethics investigation"),
        ("E-02", "First ethics complaint (Nov 2023) — MTNA took no action on conference abuse"),
        ("E-04", "TOTY 2023: no conflict disclosure; winning bio emphasized NMSM not PMTNM community service"),
        ("L-07", "Triple role conflict Nov 7, 2025: Past President/informal advisor to Sharon + Nominating Committee Chair + Parliamentarian — three undisclosed simultaneous conflicting roles; advisory channel used morning of Nov 7 to coordinate meeting strategy against VP (10:00 AM email to Sharon only: 'getting someone to the point of a motion can be challenging')"),
    ]),
    ("Marcus York (Board Member / NMSM Employee)", PURPLE_BG, PURPLE_TXT, [
        ("L-08", "Voted on financial matters (Nov 2024) while employed by Larry Blind; served on Blind-chaired Nominating Committee (Nov 2024–2025) — no formal conflict-of-interest procedure invoked; unmanaged proxy voting situation due to Blind's employment authority over York"),
    ]),
    ("Sharon Kunitz (Former President)", GOLD_BG, GOLD_TXT, [
        ("S-01", "Conflict of interest: receives scholarship money controlled by Grealish's sole sign-off"),
        ("S-02", "Reinforced financial records refusal — falsely characterized VP's duties"),
        ("S-03", "Committed on record to place agenda items on Nov 7 agenda — broke commitment"),
        ("S-04", "11-month pattern: blocked financial access, dismissed budget process, broke agenda commitment"),
        ("S-05", "Silent when Jeanne said Foundation article was 'Boring and irrelevant' — then praised it 21 days later"),
        ("M-05", "VP excluded from board packet mailing; VP's agenda items not addressed"),
    ]),
    ("Brian Shepard (MTNA CEO)", GREEN_BG, GREEN_TXT, [
        ("E-02", "Nov 2023 ethics complaint — no action"),
        ("E-01", "Jul 2024 ethics complaint re: Evan/Tatiana/NMSM — acknowledged, no action"),
        ("E-03", "Sep 2025 + Dec 2025 affiliate complaints — deferred entirely to affiliate self-governance"),
    ]),
]

ws2.cell(row=1, column=1, value="Person / Role").font = hdr_font()
ws2.cell(row=1, column=1).fill = fill(HDR_DARK)
ws2.cell(row=1, column=2, value="Violation ID").font = hdr_font()
ws2.cell(row=1, column=2).fill = fill(HDR_DARK)
ws2.cell(row=1, column=3, value="Summary").font = hdr_font()
ws2.cell(row=1, column=3).fill = fill(HDR_DARK)
for col in range(1, 4):
    ws2.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=1, column=col).border = border

current_row = 2
for person, bg, txt, violations in persons:
    for v_id, v_desc in violations:
        c1 = ws2.cell(row=current_row, column=1, value=person)
        c1.font = cell_font(color=txt, bold=True)
        c1.fill = fill(bg)
        c1.alignment = wrap()
        c1.border = border

        c2 = ws2.cell(row=current_row, column=2, value=v_id)
        c2.font = cell_font(bold=True)
        c2.fill = fill(GRAY_BG)
        c2.alignment = Alignment(horizontal="center", vertical="top")
        c2.border = border

        c3 = ws2.cell(row=current_row, column=3, value=v_desc)
        c3.font = cell_font()
        c3.fill = fill(WHITE if current_row % 2 else GRAY_BG)
        c3.alignment = wrap()
        c3.border = border

        ws2.row_dimensions[current_row].height = 36
        current_row += 1

set_col_width(ws2, 1, 30)
set_col_width(ws2, 2, 10)
set_col_width(ws2, 3, 90)

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — EMAIL THREADS INDEX
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Key Email Threads")
ws3.freeze_panes = "A2"

email_hdrs = ["Date", "From", "To / CC", "Subject", "Violation IDs", "Legal Significance"]
for col, h in enumerate(email_hdrs, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hdr_font()
    c.fill = fill(HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

email_rows = [
    ("Mar 17, 2025", "Jeanne Grealish", "Chip Miller; Sharon Kunitz", "Financial records request + refusal", "F-01", "CRITICAL — First documented financial refusal. Core basis of petition."),
    ("Aug 14, 2024", "Sharon Kunitz", "Chip Miller (VP appointment)", "PMTNM Vice-President", "—", "Establishes Chip's legitimate VP officer status from day one."),
    ("Oct 4, 2023", "Larry Blind (lblind@nmschoolofmusic.com)", "PMTNM members", "Please Vote for PMTNM Teacher of Year", "L-01", "CRITICAL — Larry uses BUSINESS email to solicit TOTY votes for his employee Tatiana."),
    ("Oct 12 – Nov 16, 2023", "Larry Blind / Tatiana Vetrinskaya / Chip Miller", "lblind@nmschoolofmusic.com; tatiana@nmschoolofmusic.com; wgmilleriii@gmail.com", "Text Message from Larry and Email from Parent (Gmail thread: 18bcf83c2e939e8c)", "E-01, L-01", "HIGH — LARRY'S UNSUBSTANTIATED ACCUSATION + REFUSAL TO PRODUCE EVIDENCE [CORRECTED FACTS]:\n\nCORRECTED FACTS: Chip was the piano technician/tuner at NMSM (not a student teacher only). Toni (morgantonid@gmail.com) is Evan's mother — Chip was teaching Evan at NMSM. Chip left NMSM partly because he identified a conflict of interest in teaching there while serving PMTNM (confirmed by Larry in his Nov 16 no-subject response: 'you decided that it was a conflict of interest to continue teaching at NMSM').\n\nTIMELINE: Oct 12 — Chip texts Larry to arrange coming in to fix a key entry problem and tune pianos (Chip is the technician). Larry replies reviewing Chip's earlier tuning work: 'Those tenths sounded pretty bad / I will do a walkthrough.' Still collegial. Oct 15 — Chip's last day at NMSM (confirmed by Tatiana in writing). Post-Oct 15 — Larry texts Chip: 'Chip, please do not contact any of your former students from our school by phone or email. This is not appropriate and I have received an email from a parent who was not very happy about your call.'\n\nCRITICAL FACT: In his Nov 14 email to Tatiana, Chip states: 'I did have a phone call to a parent at the parent's request.' The parent initiated the call. Larry's accusatory text framed this as Chip making inappropriate contact — but the parent called Chip, not the other way around.\n\nEVIDENCE WITHHELD: Chip demanded Larry produce the parent's email twice, so he could face his accusers. Neither Larry nor Tatiana ever provided it. Larry then refused to speak with Chip over the weekend. Tatiana stepped in as gatekeeper, deflecting to confirm Chip's last day (Oct 15) rather than addressing the accusation.\n\nTATIANA'S ROLE: Signs as 'Owner/Founder' — same weekend Larry is finalizing her TOTY nomination. She told Chip 'I do not see any misunderstanding or ethical violation from our side' without producing the parent email or allowing Chip to respond to the specific accusation.\n\nRESULT: Accusation left on the record. No apology from Larry. Parent's email never produced. Chip filed MTNA ethics complaint Nov 14 (E-01/E-02) as a direct consequence of this stonewalling."),
    ("Nov 16, 2023", "Chip Miller", "Tatiana Vetrinskaya; Toni (morgantonid@gmail.com — Evan's mother)", "Open communication request (Gmail thread: 18bd7f87933a6814)", "E-01", "HIGH — CHIP FOLLOWS MTNA ETHICS PROCESS CORRECTLY, OFFER NEVER ANSWERED:\n\nChip writes directly to the parent (Toni) and Tatiana citing MTNA Code of Ethics verbatim: 'The teacher shall participate in the student's change of teachers with as much communication as possible between parties, while being sensitive to the privacy rights of the student and families.'\n\nChip offers: (1) face-to-face meeting; (2) geographically neutral location of Toni's choosing; (3) professional mediator; (4) time agreeable to all; (5) pledges to stop emailing until receiving a response.\n\nThis offer was never answered. The refusal to engage in the mediation Chip proposed — despite his explicit MTNA code citation and offer of professional mediation — is what elevated the matter to a formal MTNA complaint. Chip did everything the ethics code required; the other parties did nothing."),
    ("Nov 15, 2023 (10:09 PM MT)", "Chip Miller", "lblind@nmschoolofmusic.com", "Code of Ethics Violation Alert: Two Counts (Gmail thread: 18bd4fdd71f3d45a)", "E-01, E-02", "HIGH — CHIP FORMALLY CHARGES LARRY WITH TWO MTNA CODE VIOLATIONS:\n\nCOUNT 1 (Commitment to Colleagues): Larry's accusatory text — reprimanding Chip for calling Evan's mother (Toni), a call Toni herself initiated about Evan's student transfer.\n\nCOUNT 2 (Commitment to Society): Larry refused to speak with Chip at the PMTNM conference on Nov 10 AND Nov 11. Chip documents the verbatim exchanges: Nov 10 — Chip: 'Can I talk with you?' Larry: 'No.' [walks away]. Nov 11 — same immediate 'No,' then 'I have to get the programs typed.' Chip offered to help; Larry refused that too. Chip notes: 'I could have been of great help to you, given the large number of errors that your program eventually contained.' This conference = same event where NMSM students were featured (E-02) and TOTY vote was active with Tatiana as Larry's nominee.\n\nForwarded to MTNA Exec. Dir. Gary Ingle Nov 22, and again Dec 7 noting Larry's silence since Nov 16. Gary Ingle took no action (see E-02)."),
    ("Nov 15, 2023 (8:17 PM MT)", "Larry Blind (lblind@nmschoolofmusic.com)", "wgmilleriii@gmail.com", "(no subject) — Larry's only written defense (Gmail thread: 18bd6226826c5972)", "E-01, E-02, L-04", "CRITICAL — LARRY'S ADMISSIONS ON RECORD (his only written response to the ethics counts):\n\nADMISSION 1 — Conference refusals confirmed: 'I politely told you no. I believe my words were No Chip, I can't, not right now.' Does not address why he could not spare 5 seconds to ask what the matter was about.\n\nADMISSION 2 — Parent email withheld: 'the parent in question emailed me, not you. The email covered several topics and it was not meant to be shared or it would have included other recipients. I am not required by any ethics rules to share the email with you.' KEY: Larry received Toni's email, used it to accuse Chip of inappropriate contact via text, then refused to let Chip see the accusation. Chip was reprimanded on evidence he was never allowed to see. Toni had called Chip at her own request.\n\nADMISSION 3 — Chip identified conflict of interest: 'you decided that it was a conflict of interest to continue teaching at NMSM, even after both Tatiana and I explained to you that we did not see it that way.' Chip's ethical awareness of the NMSM/PMTNM conflict predates his VP role by a year. Larry and Tatiana wanted him to overlook it and keep teaching.\n\nSHUTDOWN: 'This will be my last communication with you about this matter.' Same pattern as PMTNM 2025: accusation made, substance refused, process cut off unilaterally."  ),
    ("Nov 14, 2023", "Chip Miller", "Gary Ingle (MTNA Exec. Dir.)", "Ethical Dispute assistance request", "E-02", "First MTNA ethics complaint re: NMSM students at PMTNM conference."),
    ("Jul 27 – Aug 9, 2024", "Chip Miller / Brian Shepard", "Brian Shepard (MTNA CEO)", "Ethical Dispute Mediation Request", "E-01", "Ethics complaint re: Evan/Tatiana/NMSM. Brian reviews but backs off."),
    ("Sep 27–28, 2024", "Haewon Yang / Chip Miller", "Chip Miller / Haewon Yang", "FTC guidelines for MTNA and PMTNM", "L-01, E-01", "Haewon confirms TOTY nomination process; 'code of ethics cannot restrict membership'; NMSM handbook blocks complaint."),
    ("Aug 31 – Sep 16, 2024", "Haewon Yang / Chip", "Chip Miller", "Moving forward with PMTNM", "E-01, L-01", "Chip: NMSM teaching contract breach referenced. Haewon tries to mediate."),
    ("Sep 28, 2024", "Chip Miller / Jacqueline Zander-Wall", "zanderwall@gmail.com", "Fwd: Please Vote for PMTNM Teacher of Year", "L-01", "Chip investigates TOTY records; Zander-Wall declines to provide them."),
    ("Dec 2024 / Jan 2025", "Larry Blind", "Jeanne Grealish", "Bank account removal request", "F-05", "CRITICAL — Larry warns Grealish should NOT be sole bank signatory. Warning ignored."),
    ("Jan 2025", "Sharon Kunitz; Jeanne Grealish", "Chip Miller; Larry Blind", "Handbook threads (multiple)", "R-01", "Grealish falsely claims handbook is online. Chip cannot find it."),
    ("Oct 2, 2025", "Chip Miller", "Sharon, Grealish, Larry, Laura", "Re: Foundation (budget committee request)", "F-02", "Chip formally asks for budget committee. Grealish will later refuse."),
    ("Oct 14, 2025", "Chip Miller", "Sharon Kunitz", "ITEMS FOR BOARD AGENDA (Nov 7-8)", "S-03, F-02", "CRITICAL — Formal 5-item agenda request. Sharon dismissively replies."),
    ("Oct 14, 2024 (approx)", "Larry Blind", "PMTNM Board + General Membership", "PMTNM Board and General Membership Meetings — cell phone voting announcement (Gmail thread: 1928ce26b2fe3d9c)", "G-02", "HIGH — 2024 ANONYMOUS VOTING PRECEDENT: Larry announced that online members wishing to cast 'anonymous' secret ballot votes should text to his personal cell phone (505) 980-3941. Meeting held at NM School of Music — Larry's own commercial school.\n\nWHY THIS IS NOT ANONYMOUS: Larry, as the sole recipient, could see every vote cast — identical in structure to his Nov 7, 2025 counter-proposal ('designate one person to receive the texts, count the results, and then delete the texts') that Chip rejected as non-anonymous under RROO §§45:5–7, 46:32–34. A vote-receiver who sees all votes before counting cannot provide the anonymity Robert's Rules requires for ballot votes.\n\nPATTERN ACROSS TWO CONSECUTIVE ELECTIONS: 2024 — votes texted to Larry's personal cell at his own school. 2025 — Chip built a proper anonymous system (ElectionChamp, 15 board members loaded by 10:55 AM); Larry rejected it that morning and ran the no-confidence vote publicly instead with Grealish in the room. Two election cycles, zero genuine anonymous voting available to online members."),
    ("Oct 25–30, 2025", "Jeanne Grealish", "Board (Larry, Chip, Laura)", "PMTNM Budget / FY Report", "F-03", "CPA report 5 months late. Budget confusion — Grealish unsure which year."),
    ("Oct 31–Nov 6, 2025", "Jeanne Grealish; Larry Blind; Sharon Kunitz; Chip Miller", "Grealish, Chip, Larry, Sharon, Heather Nasi, Laura Spitzer; CC to Kathy Black and Karen Kilgore (Grealish's attorney)", "Budget Committee REMINDER — Email-based budget approval & pre-litigation coordination (Gmail thread: 19a3c6e590c51011)", "F-02, F-17, G-02, M-07", "CRITICAL — FOUR DISTINCT VIOLATIONS IN SINGLE THREAD:\n\n(1) EMAIL-BASED BUDGET PROCESS BYPASSES HANDBOOK: Grealish: 'This method of arriving at a proposed budget has worked well for several years' — no formal Budget Committee meeting held. Handbook (Art. IX §2) requires VP as committee member; handbook (Section VI-B) requires formal Executive Board meeting approval. Neither occurred. Oct 31–Nov 6 exchange is purely email.\n\n(2) LARRY'S RUBBER-STAMP APPROVAL (F-02, F-17): Nov 2, 2025 — Larry replies 'The budget looks ok to me' to the thread, explicitly saying questions 'do not alter the basis for approving the budget as presented.' This is neither a motion nor formal approval — it's a casual reply-all email from the Past President about a budget containing Grealish's $1,750 honorarium (F-17).\n\n(3) CHIP'S UNANSWERED QUESTIONS ABOUT HONORARIUM COMPENSATION (F-17): Oct 31 — Chip formally asks: 'Has your honorarium increased through the years as our membership has decreased? What are the measures for your honorarium?' Sharon's response is dismissive: 'Are you a CPA or Finance expert? You have not been a member long enough to know the operations of PMTNM.' The questions were never answered. The honorarium was approved (falsely per M-07) one week later without addressing Chip's concerns.\n\n(4) ESCALATION TO ATTORNEYS TRIGGERS COORDINATION (G-02): Nov 5, 7:03 PM — Grealish forwards Chip's retirement request email to Brian Shepard (MTNA national) and Karen Kilgore (Grealish's attorney). This event preceded the Nov 6 pre-meeting coordination. Hours later, Nov 6 morning emails show Larry rejecting anonymous voting and Sharon setting up meeting structure (pre-litigation coordination documented in G-02).\n\n(5) KATHY BLACK DISCOVERY: Nov 6 — Grealish had forwarded the entire thread to both MTNA and her attorney. Kathy Black (Chip's counsel) learned from Chip that he was being opposed by coordinated parties and that Grealish had already retained legal counsel and looped in MTNA.\n\nRESULT: Budget 'approved' via email with no formal motion or vote. Nov 7 minutes falsified the approval (M-07). Officer compensation (F-17) authorized without genuine board action. Anonymous voting suppressed on Nov 6 morning in response to Nov 5 escalation (G-02)."),
    ("Oct 31, 2025", "Jeanne Grealish", "Board", "Budget Committee REMINDER (condensed)", "F-02", "Chip asks formal budget questions — most ignored. Larry witnesses."),
    ("Oct 31, 2025", "Sharon Kunitz", "Kathy Black (cc: Board)", "PMTNM position", "S-02", "Sharon disclaims PMTNM involvement in dispute — written to Chip's attorney."),
    ("Nov 4, 2025 1:14:51", "Sharon Kunitz (on recording)", "Meeting attendees", "Nov 4 Finance Meeting — agenda commitment", "S-03", "CRITICAL — Sharon says 'I'll put them on the agenda.' Confirmed by Kathy Black email same day at 2:54 PM."),
    ("Nov 4, 2025 2:54 PM", "Kathy Black", "Chip Miller", "Follow-Up to Today's Meeting", "S-03, E-01", "CRITICAL — Attorney confirms Sharon's agenda commitment + Grealish attorney surprise appearance."),
    ("Nov 5–6, 2025", "Larry Blind", "Chip Miller (cc: Sharon, Jeanne, Laura)", "Willis Glen Miller, III", "L-04", "CRITICAL — Larry hostile email night before board meeting. Then serves as Parliamentarian."),
    ("Nov 6, 2025 (11:40 AM – 10:15 PM)", "Cheryl Pachak-Brooks → Chip Miller", "wgmilleriii@gmail.com", "PMTNM concerns (Gmail thread: 19a5a7881ce060db)", "L-04, S-03", "HIGH — PRE-MEETING CHARACTER COORDINATION: Cheryl Pachak-Brooks (former VP, no current board role) contacted Chip the day before the Nov 7 board meeting, having been briefed by an undisclosed party. Cheryl's opening: 'I have been made aware of your issues with PMTNM and Jeanne' and 'When I heard about everything that has transpired, my first thought was, he is trying to dismantle the association. I really regretted suggesting you get involved.' Her language (attacking, cruel, dismantling) mirrors exactly the characterizations used by Grealish/Sharon/Larry — indicating she was briefed by one or more of them before reaching out.\n\nChip offered twice to share his documentation and offered a phone call. Cheryl declined both: 'I have heard several sides and been apprised of all that is going on. I know there are two sides, I am just suggesting you step above the fray.' She formed a conclusion without hearing his evidence.\n\nThis contact occurred the same day as: (1) Larry's hostile email to Chip (L-04); (2) Sharon's Robert's Rules email setting up meeting structure (M-01); (3) The pre-meeting anonymous voting exchange (G-02). Cheryl then attended Nov 7 on Zoom.\n\nCHIP'S FINAL REPLY — a list of 19 values he called his heart: Diplomatic / Kind / A thriving organization / Avoidance of lawsuits / No meanness / No name calling / Respect / Dignity / Patience / The continuation of peace / The continuation of PMTNM / Friendship / Cooperation / Music making / Organization / Legal compliance / Responsibility with our money / Responsible with each others money / Open communication. This is the clearest single statement of what Chip was actually working toward — written the night before the meeting that dismantled his VP role."),
    ("Nov 7, 2025 7:07 AM", "Sharon Kunitz", "Board (Larry, Laura)", "Additional RRoo wisdom!", "M-01", "CRITICAL — Sharon's own email establishes 15-min rule. Minutes later say 10 min."),
    ("Nov 7, 2025 (pre-meeting, 7:07–10:55 AM)", "Sharon Kunitz / Chip Miller / Larry Blind", "All officers", "Additional RRoo wisdom! — Anonymous voting exchange (Gmail thread: 19a5ea6717e3582c)", "L-04, G-02", "CRITICAL — Chip set up ElectionChamp anonymous voting (RROO §§45:5–7, 46:32–34 compliant) and sent test election to all 15 board member emails by 10:55 AM. Larry counter-proposed text-to-one-phone (not anonymous — one person sees all votes). Chip explained why this fails RROO ballot standards. Larry\'s 10:00 AM reply to Sharon (ignoring Chip\'s proposal): \'There may be a lot of talking, but getting someone to the point of a motion can be challenging\' — hours before the meeting, telegraphing his plan to prevent the motion procedurally. System was ready and never used. Vote of no confidence run as public hand-raise instead."),
        ("Nov 7, 2025 (meeting)", "Audio recording", "Full board", "Nov 7 Board Meeting", "M-01 through M-05, F-04, L-04, G-02", "CRITICAL — Full audio record. Contradicts minutes on 8 documented points. Key sequences: (1) Chip: 'I have no packet' [32:28] — excluded from board materials. (2) Grealish identifies VP report but says 'It cannot be board reports, because I don\'t have anything to include' [33:18]. (3) Larry copies report, then questions authorship as delay tactic [34:11]. (4) Sharon attempts adjournment without covering Chip\'s agenda items [41:30]. (5) Voice: 'You didn\'t send that' — Sharon immediately contradicts: 'two months ago, March 17, thank you' [42:43–42:52]. (6) Chip makes formal motion of no confidence [46:34]. (7) Larry controls second process, no one seconds, motion dies [48:04–49:00]. (8) Sharon adjourns [49:13]. Minutes: 'Old Business: None. New Business: None.'"),
    ("Sep 23 + Dec 23, 2025", "Chip Miller", "Brian Shepard (MTNA CEO)", "PMTNM Ethics/Compliance Complaints", "E-03", "Third and fourth complaints to MTNA. Shepard defers to affiliate self-governance."),
    ("Mar 9, 2026", "Jeanne Grealish", "Full board (Larry, Laura, Chip)", "PMTNM MINUTES", "R-02", "Minutes distributed 4 months late — after petition filed, after response filed."),
    ("Mar 13, 2026", "Chip Miller", "Laura Spitzer", "November 7 Board Meeting Minutes", "M-01 through M-05", "Chip formally raises minutes inaccuracies with new president. No substantive reply."),
    ("Sep 11, 2025", "Jeanne Grealish", "admin@pmtnm.org (Chip)", "Re: PMTNM | MTNA | The Foundation", "S-05", "CRITICAL — Grealish: 'Boring and irrelevant to most of us.' Response to MTNA Foundation newsletter. Dismissive tone toward MTNA's core charitable mission."),
    ("Sep 12, 2025 12:44 AM", "Chip Miller", "Grealish + Brian Shepard + Sharon Kunitz", "Fwd: PMTNM | MTNA | The Foundation", "S-05, E-03", "CRITICAL — Chip forwards Jeanne's 'Boring and irrelevant' email to Sharon, asking 'Hi Jeanne, can you clarify?' Sharon was explicitly CC'd. Sharon said nothing."),
    ("Oct 2, 2025", "Sharon Kunitz", "Chip Miller (CC: Jeanne, Larry, Laura)", "Foundation", "S-05", "CRITICAL CONTRAST — Sharon: 'Great article on Foundation! Beautifully written and presented.' — 21 days after receiving Jeanne's 'Boring and irrelevant' email and saying nothing."),
    ("Oct 7, 2025", "Jeanne Grealish → Chip, Sharon, Laura, Larry, Heather Nasi", "All officers", "Re: Foundation. RE: BUDGET — 'No legal mandates' email (Gmail thread: 199c0d60237afa6c)", "F-02, G-01, F-14", "CRITICAL — THE FULL CONTEXT: Chip had asked about the budget committee meeting. Grealish responded: 'There probably will be no need for an actual meeting.' Then: 'As for any statements in the Handbook regarding the budget, committee, dates, etc. these would be grossly out of date since the Handbook itself is no longer current and requires serious revision. Thanks heaven, PMTNM does not function under any legal mandates!' This was a direct, specific rejection of the handbook's budget committee provisions — not a philosophical aside. ALL FIVE OFFICERS received this email."),
    ("Oct 23, 2025", "Kathy Black (Atty.) → Chip Miller", "Chip Miller", "Initial Impressions — NM Secretary of State annual report overdue since 2019", "F-16, F-05", "CRITICAL — Kathy Black: 'according to New Mexico Secretary of State business records, PMTNM's annual report is way overdue, with the due date listed as November 15, 2019.' Found within days of beginning review of PMTNM records. Also noted: 'You've done a tremendous amount of work in a short period of time. Impressive.' — attorney acknowledgment of Chip's documentation effort. PMTNM's own financial documents state 'All federal, state and local required documents and lists have been prepared and filed' — a direct false statement contradicted by the Secretary of State public record."),
    ("Oct 29, 2025", "Kathy Black (Atty.) → Grealish; Grealish private letter → Kathy Black", "Forwarded to Chip Oct 30", "Attorney demand + Grealish private response (Gmail thread: 19a3309ed58127ac)", "F-01, A-01, M-06, F-05", "CRITICAL — Kathy Black formally demanded records under NMSA §53-8-27 by Oct 31. Grealish's private response (forwarded to Chip by Kathy): 'It is not physically possible for me to bring the requested items to your office on Friday.' 'PMTNM does not do online banking.' 'PMTNM does not have a separate Finance Committee. It has not had an Auditing Comm. since hiring a CPA.' 'Past records and tax documents (both federal and state) are kept securely in storage along with Minutes from Board and General Meetings.' 'Current files...bank statements for all seven accounts...are kept in my home office at 1226 Morningside Dr. NE.' This single letter is Grealish's own written description of PMTNM's governance structure — sole home control, no audit committee, minutes in storage, no online access. Attorney-witnessed."),
    ("Oct 29–31, 2025", "Jeanne Grealish → Budget Committee", "Sharon, Larry, Chip, Heather Nasi (CC: Laura)", "PMTNM FY report and Proposed Budget for 2025-2026 (Gmail thread: 19a3309ed58127ac)", "F-02, F-03, F-14", "CRITICAL — Budget and CPA report sent 4 months into fiscal year. Grealish: 'In recent years instead of meeting we have worked via email since there were few adjustments to the budget draft.' Kathy Black: 'the organization is not following the handbook regarding committees' and 'the State of New Mexico annual report has not been filed.' Bank balances as of Oct 29: General Fund $17,475.18, Money Market $9,719.58, Joyce Walker $23,942.38, Student Travel $3,086.85, Danfelser $2,707.14, Weed $467.55, Composition $939.76, Jane Snow $2,346.73. Total: $60,685.17 + Danfelser CD $3,969.51."),
    ("2025 (TOTY cycle)", "Jeanne Grealish", "Board", "2025 TOTY Nomination — Terri Reck", "F-11", "HIGH — Grealish made 3 attempts to nullify Terri Reck's TOTY nomination. Stated Reck 'has not contributed very much to community service.' Immediately contradicted by board members who seconded Reck."),
    ("Pre-Nov 7, 2025", "Chip Miller", "President + EST", "VP Report Submitted to President and EST", "M-04, G-01, F-11", "CRITICAL — VP Report submitted in advance of Nov 7 meeting. Directly contradicts minutes claim that 'VP report had not been submitted.' Preserved on Google Drive (ID: 1d-vg5v87N7DwF7EUisJ6M-HYiYrbuC3Z)."),
    ("Sep 30, 2023", "Lawrence Blind (lawrenceblind@msn.com)", "PMTNM (submitted as TOTY nomination)", "TOTY Nomination — Tatiana Vetrinskaya [PHYSICAL PDF REVIEWED]", "L-05, E-04", "CRITICAL — Larry submitted Tatiana's nomination from his PERSONAL email (not NMSM employer email, unlike his vote solicitation). Tatiana's listed email is tatiana@nmschoolofmusic.com. Community service section = entirely NMSM activities. Future-tense item listed as achievement: 'will present a workshop at 2023 PMTNM Conference.' Closing attribute: 'business acumen — it is what I know she is most proud of.' This phrase ('I know she is most proud of') reveals Larry wrote it — a nominee does not describe herself this way."),
    ("Sep 30, 2023", "Jacqueline Zander-Wall (zanderwall@gmail.com)", "PMTNM (submitted as TOTY nomination)", "TOTY Nomination — Terri Reck [PHYSICAL PDF REVIEWED]", "L-05, E-04, F-11", "HIGH — Zander-Wall submitted Reck's nomination. Terri Reck: 47 years teaching (1976–2023), Dalcroze certification across 6 institutions, 5 AMTA committee leadership roles, harpsichordist with SF Symphony / NM Symphony / NM Philharmonic, students in PMTNM/MTNA programs, Music Bowl Chair for decades, jewelry donated to PMTNM silent auctions. Every criterion directly met. This is the nomination that lost to Larry's 'storybook language' submission."),
    ("Tax Year 2018", "NM Attorney General", "PMTNM (FEIN 85-0284938)", "NM AG Charitable Organization Registration Tax Year 2018 [PHYSICAL DOCUMENT REVIEWED]", "F-05, F-12", "HIGH — 2018 filing: Grealish sole responsible party for all 5 financial functions; PMTNM address = personal home; NTEE = A6B (misclassification); FY2018 revenue $18,817 vs. expenses $22,649 (deficit); 'Solicited funds in NM: No' (appears inaccurate). See also 2023 filing entry below."),

    ("Oct 23, 2025", "Jeanne Grealish (executing for PMTNM)", "VIP Staffing", "VIP Staffing Master Terms & Conditions — Executed [PHYSICAL CONTRACT REVIEWED]", "F-13", "HIGH — Commercial staffing contract executed by Grealish for PMTNM on Oct 23, 2025. Three copies in Drive (draft + executed + third copy). Requires PMTNM to carry GL and EPLI insurance. Contains 1-year non-solicitation clause binding on PMTNM. No board authorization in any minutes. Nov 7 board meeting (17 days later) contains no reference to this contract."),
]

email_widths = [18, 25, 25, 35, 15, 50]
for col, w in enumerate(email_widths, 1):
    set_col_width(ws3, col, w)

for i, row in enumerate(email_rows, 2):
    flag = row[4] != "—" and ("CRITICAL" in row[5] or row[4] in ["S-03", "L-01", "L-04", "M-01", "M-02", "F-01", "F-04"])
    bg = RED_BG if flag else (GRAY_BG if i % 2 == 0 else WHITE)
    for j, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = cell_font(color=RED_TXT if flag else "000000", bold=flag and j in [4, 5])
        c.fill = fill(bg)
        c.alignment = wrap()
        c.border = border
    ws3.row_dimensions[i].height = 54

ws3.row_dimensions[1].height = 30

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — MTNA CODE OF ETHICS CROSS-REFERENCE
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("MTNA Ethics Cross-Ref")
ws4.freeze_panes = "A2"

# Header
mtna_headers = ["MTNA Principle", "Code Section", "Key Quote / Standard", "Violations That Apply", "Notes"]
mtna_col_widths = [30, 16, 55, 28, 40]
for col, h in enumerate(mtna_headers, 1):
    c = ws4.cell(row=1, column=col, value=h)
    c.font = hdr_font()
    c.fill = fill(HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws4.column_dimensions[get_column_letter(col)].width = mtna_col_widths[col-1]

# MTNA principles mapped to violations
SLATE_FILL = PatternFill("solid", start_color="D6DCE4", fgColor="D6DCE4")
TEAL_FILL  = PatternFill("solid", start_color="D9EAD3", fgColor="D9EAD3")
ROSE_FILL  = PatternFill("solid", start_color="FCE4D6", fgColor="FCE4D6")
GOLD_FILL  = PatternFill("solid", start_color="FFF2CC", fgColor="FFF2CC")
BLUE_FILL  = PatternFill("solid", start_color="DDEEFF", fgColor="DDEEFF")

severity_fills = {
    "CRITICAL": ROSE_FILL,
    "HIGH":     GOLD_FILL,
    "MEDIUM":   BLUE_FILL,
    "LOW":      SLATE_FILL,
}

mtna_rows = [
    ("Honest Representation of Our Art",
     "Preamble / §1",
     "'Music teachers shall represent their professional qualifications, experience, and accomplishments honestly.'",
     "L-06 (CRITICAL)\nL-05, E-04 (HIGH)",
     "Larry's NMSM biography: 'perform internationally' for paid European recitals; Carnegie Hall framing without substantiation; uncredentialed faculty presented as experienced teachers. Directly violates MTNA's foundational principle.",
     "HIGH"),

    ("Conflict of Interest Disclosure",
     "§3",
     "'Members shall disclose and manage conflicts of interest that could compromise their professional judgment.'",
     "L-01, L-05 (CRITICAL)\nE-04, S-01 (HIGH)",
     "Larry administered the TOTY process and voted in his own employee (Tatiana). No disclosure. Sharon receives scholarship disbursements controlled by Grealish and consistently votes with Grealish. Neither conflict was ever disclosed.",
     "CRITICAL"),

    ("Student Welfare and Transfer Protocols",
     "§4",
     "'Members shall protect the welfare of students during transitions between teachers.'",
     "E-01 (HIGH)\nE-04 (HIGH)",
     "NMSM employee handbook invoked to block Chip's communication with Evan's mother during a student transfer. Tatiana declined to meet with Evan's parent to clear the air. Larry silent throughout. Brian Shepard used NMSM's internal policies as a shield.",
     "HIGH"),

    ("Professional Conduct Among Colleagues",
     "§5",
     "'Members shall treat colleagues with respect and refrain from disparagement.'",
     "G-01 (CRITICAL)\nS-05, F-08 (HIGH)",
     "Grealish: 'PMTNM does not function under any legal mandates!' (Oct 7 EST). 'Waste of everyone's time.' 'Senseless cluttering up of mailboxes.' Grealish stated Teri Reck 'hadn't contributed.' Jeanne: 'Boring and irrelevant to most of us' re: MTNA Foundation. Sharon silent throughout.",
     "CRITICAL"),

    ("Fair and Transparent Organizational Processes",
     "§6",
     "'Members serving in leadership roles shall ensure fair, transparent, and inclusive governance processes.'",
     "M-01 through M-05 (CRITICAL)\nF-01, F-04, S-03, S-04 (HIGH)",
     "Minutes falsified on 5 documented points. Financial report 'approved' with no vote. VP excluded from board packet. Agenda commitment broken. VP denied financial records for 11+ months. TOTY nomination extended unilaterally.",
     "CRITICAL"),

    ("Financial Integrity and Accountability",
     "§7",
     "'Members in fiduciary roles shall maintain rigorous financial accountability and transparency.'",
     "F-01, F-04, F-05 (CRITICAL)\nF-06, F-10, F-11 (HIGH)",
     "Sole signatory on all accounts. Personal furniture in org storage (attorney-witnessed). 53% storage budget overrun. Financial report approved without vote. CPA report 5 months late. Budget adopted without committee process. Grealish's own statement: 'PMTNM does not function under any legal mandates.'",
     "CRITICAL"),

    ("Compliance with Law and Affiliate Obligations",
     "§8",
     "'Members shall comply with applicable law and fulfill obligations to their MTNA affiliate.'",
     "G-01 (CRITICAL)\nF-01, R-01 (HIGH)",
     "Grealish explicitly repudiated legal mandates on the record. Handbook withheld 6 months. Financial records refused in violation of NM Nonprofit statute. MTNA filed three complaints — zero action taken by Brian Shepard.",
     "CRITICAL"),

    ("Institutional vs. Individual Membership (FTC Context)",
     "FTC revision (2014)",
     "Code revised with FTC attorneys to prevent antitrust issues — cannot restrict institutional membership.",
     "E-01, L-01 (CRITICAL)\nL-02, L-06 (HIGH)",
     "PARADOX: FTC revision meant to protect teacher rights now shields NMSM's institutional participation in PMTNM. Brian Shepard: excluding NMSM 'could be considered an anticompetitive practice.' NMSM's enrollment growth directly harms PMTNM member teachers' independent studios.",
     "HIGH"),

    ("Confidentiality and Discretion",
     "§9",
     "'Members shall exercise discretion regarding confidential information about students, families, and colleagues.'",
     "E-01 (HIGH)",
     "Larry allegedly spoke negatively to Evan's mother about Chip, damaging Chip's professional relationship with the Menaul School. Details of the student transfer were disclosed in ways that harmed all parties.",
     "HIGH"),
]

# Write MTNA rows
for i, (principle, section, quote, violations, notes, severity) in enumerate(mtna_rows, 2):
    row_fill = severity_fills.get(severity, SLATE_FILL)
    vals = [principle, section, quote, violations, notes]
    for j, val in enumerate(vals, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.fill = row_fill if j == 1 else (PatternFill("solid", start_color="F9F9F9", fgColor="F9F9F9") if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", fgColor="FFFFFF"))
        c.font = cell_font(bold=(j == 1))
        c.alignment = wrap()
        c.border = border
    ws4.row_dimensions[i].height = 90

ws4.row_dimensions[1].height = 30

# ── KEY QUOTES BLOCK ─────────────────────────────────────────────────────────
# Add a section below with verbatim quotes for courtroom/hearing use
quote_start = len(mtna_rows) + 3
ws4.cell(row=quote_start, column=1, value="KEY VERBATIM QUOTES — DOCUMENTARY EVIDENCE").font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws4.cell(row=quote_start, column=1).fill = fill(HDR_MED)
ws4.merge_cells(start_row=quote_start, start_column=1, end_row=quote_start, end_column=5)
ws4.cell(row=quote_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[quote_start].height = 24

key_quotes = [
    ("Jeanne Grealish — Oct 7, 2025 EST Meeting",
     "'PMTNM does not function under any legal mandates!'",
     "G-01", "CRITICAL",
     "On-record repudiation of 501(c)(3) obligations. Source: VP Report Section 3.3 (Drive ID: 1d-vg5v87N7DwF7EUisJ6M-HYiYrbuC3Z)"),
    ("Jeanne Grealish — Re: Foundation newsletter",
     "'Boring and irrelevant to most of us.'",
     "S-05", "HIGH",
     "Response to MTNA Foundation newsletter. Sharon CC'd on clarification request and said nothing. Sharon then: 'Great article! Beautifully written and presented' 21 days later."),
    ("Haewon Yang — Sep 28, 2024 (re: 2023 TOTY)",
     "'I voted for the person who had a bio that sounded like they deserved the award.'",
     "L-05, E-04", "CRITICAL",
     "Admission that the TOTY nomination was won by presentation of a bio — not by merit evaluation. Larry controlled the process and the winner was his employee."),
    ("Larry Blind — Nov 2, 2025 email",
     "[Re: storage unit] raising the $4,600 operating expense question.",
     "F-10", "HIGH",
     "Even the person whose employer relationship benefits from Grealish's goodwill questioned the storage expense in writing. Grealish had already admitted personal furniture in the unit before attorney Kathy Black."),
    ("Brian Shepard — MTNA CEO, Ethics complaint thread",
     "'This is the first time we've had an issue that wasn't between two independent teachers.'",
     "E-01", "HIGH",
     "MTNA's own CEO acknowledging the Code had never contemplated an institution vs. individual teacher dispute. FTC revision created the gap Shepard then used to avoid action."),
    ("Carnegie Hall / NMSM framing note",
     "'Students from the NMSM have an opportunity to perform internationally.'",
     "L-06", "HIGH",
     "NMSM students buy tickets and fly to Europe to play in recitals — this is not 'performing internationally' in any professional sense. The framing misleads prospective families and devalues what true international performance achievement means."),
    ("Sharon Kunitz — Oct 2, 2025",
     "'Great article on Foundation! Beautifully written and presented.'",
     "S-05", "HIGH",
     "Sent 21 days after receiving Jeanne's 'Boring and irrelevant' response — CCed to Jeanne, Larry, and Laura. Sharon knew. She chose not to intervene in the moment."),
    ("Larry Blind — TOTY Nomination for Tatiana Vetrinskaya (Sep 30, 2023)",
     "'...business acumen... it is what I know she is most proud of.'",
     "L-05, E-04", "CRITICAL",
     "Closing line of the winning TOTY nomination. 'I know she is most proud of' is written from Larry's perspective — proving he authored the document. 'Business acumen' is not a PMTNM TOTY award criterion. Physical PDF reviewed; submitted from lawrenceblind@msn.com."),
    ("Larry Blind — TOTY Nomination for Tatiana Vetrinskaya (Sep 30, 2023)",
     "'[Tatiana] will present a workshop at 2023 PMTNM Conference.'",
     "L-05, E-04", "CRITICAL",
     "Future-tense item listed as an achievement in a nomination submitted Sep 30, 2023. The PMTNM Conference had not yet occurred. Listing a future event as a credential is a substantive misrepresentation of the nominee's record at the time of submission."),
    ("NM AG Registration — PMTNM (Tax Year 2018)",
     "Solicited funds in NM: No. [All 5 financial functions: Grealish. Address: personal home.]",
     "F-05, F-12", "HIGH",
     "State-filed document lists PMTNM's registered address as Grealish's personal residence, designates her as sole responsible party for all financial functions, and declares the organization did not solicit funds in NM — despite collecting dues, conference fees, and Foundation contributions. Physical document reviewed: 20184721935751263.pdf."),
    ("Jeanne Grealish — Oct 7, 2025 email (re: Budget Committee)",
     "'There probably will be no need for an actual meeting.'",
     "F-02, F-14", "CRITICAL",
     "Said in direct response to VP's request for a budget committee meeting — a meeting mandated by the PMTNM Handbook (Bylaws Art. IX §2). Followed immediately by: 'Thanks heaven, PMTNM does not function under any legal mandates!' The two statements together show deliberate dismissal of a specific governance requirement. Gmail thread ID: 199c0d60237afa6c. All five officers CC'd."),
    ("Jeanne Grealish — Oct 7, 2025 email, same message",
     "'As for any statements in the Handbook regarding the budget, committee, dates, etc. these would be grossly out of date since the Handbook itself is no longer current.'",
     "F-02, G-01", "CRITICAL",
     "Grealish preemptively dismissed the authority of the handbook — the organization's own governing document — to avoid the budget committee obligation. Immediately followed by: 'Thanks heaven, PMTNM does not function under any legal mandates!' The handbook is available online at pmtnm.org/handbook_view/view_section.php and contains binding bylaw provisions."),
    ("Kathy Black (Chip's attorney) — Oct 30, 2025",
     "'The organization is not following the handbook regarding committees.'",
     "F-02, F-14", "HIGH",
     "Attorney's independent legal assessment after reviewing the budget email exchange. Also noted: 'At the very least, I am going to point out that the State of New Mexico annual report has not been filed.' Attorney-verified conclusion about handbook non-compliance."),
    ("Jeanne Grealish — Private letter to Kathy Black, Oct 30, 2025",
     "'PMTNM does not have a separate Finance Committee. It has not had an Auditing Comm. since hiring a CPA to double check all activity and file all required reports with IRS and the State of NM.'",
     "A-01", "CRITICAL",
     "Written admission to VP's attorney that two standing committees required by the PMTNM Handbook have been unilaterally eliminated. Forwarded to Chip by Kathy Black (Gmail thread ID: 19a3309ed58127ac). Kathy Black confirmed: 'The function of an audit committee should be to retain a CPA for a formal audit, which this organization does not do.'"),
    ("Jeanne Grealish — Private letter to Kathy Black, Oct 30, 2025",
     "'Past records and tax documents (both federal and state) are kept securely in storage along with Minutes from Board and General Meetings.'",
     "M-06, F-05", "CRITICAL",
     "Board meeting minutes — PMTNM's legal record of all corporate decisions — are in the storage unit. The same unit where Grealish stores personal furniture (F-06, attorney-witnessed). Current financial records are at her home. No online banking. When attorney demanded records, Grealish said it was 'not physically possible' to comply by the statutory deadline."),
    ("Jeanne Grealish — Private letter to Kathy Black, Oct 30, 2025",
     "'It is not physically possible for me to bring the requested items to your office on Friday, October 31, 2025. PMTNM does not do online banking.'",
     "F-01, M-06", "CRITICAL",
     "Response to a formal attorney demand letter citing NMSA §53-8-27. Grealish refused production to VP's attorney within the statutory period, claiming physical impossibility. This is the same refusal pattern documented in March 2025 (F-01) — now repeated against an attorney's formal demand. 'PMTNM does not do online banking' confirms no alternative production method exists."),
    ("2004 Handbook Section VII-D — Auditing Committee",
     "'Makes a complete audit of the books prior to (if possible) or during the State Conference. All figures in the ledgers, journals, check book, receipts, and vouchers should be checked.'",
     "A-01", "HIGH",
     "The handbook requires this independent internal audit annually at the State Conference. Grealish eliminated it, substituting a CPA who prepares tax returns based solely on materials Grealish provides. No independent verification of ledgers, journals, checkbook, or vouchers has occurred since the Auditing Committee was disbanded."),
    ("PMTNM internal financial document — 'Budget Documents, drafts approved by Board (1).pdf'",
     "'All federal, state and local required documents and lists have been prepared and filed.'",
     "F-16", "CRITICAL",
     "This statement appears in a PMTNM document filed with or presented to the board. It is directly contradicted by the NM Secretary of State public record showing the annual report overdue since November 15, 2019. If Grealish presented or approved this language, she made a false representation to the board. Source: Google Drive PMTNM folder, 'Budget Documents, drafts approved by Board (1).pdf'."),
    ("Larry Blind — Nov 7, 2025, 10:00 AM, pre-meeting email (Gmail: 19a5ea6717e3582c)",
     "'There may be a lot of talking, but getting someone to the point of a motion can be challenging.'",
     "L-04, G-02", "CRITICAL",
     "Written hours before the board meeting — in reply to Sharon\'s Robert\'s Rules email, ignoring Chip\'s anonymous voting proposal entirely. This is not a procedural observation; it is a preview of strategy. Larry served as Parliamentarian at the meeting that afternoon and controlled the second process for Chip\'s motion of no confidence, calling for a hand-raise in a room with Grealish present, without using the anonymous voting system Chip had set up and ready by 10:55 AM. The motion died without a second. The quote proves the outcome was anticipated and planned."),
        ("Sharon Kunitz — Oct 14, 2025 reply to Chip's formal 5-item agenda request (Gmail: 199e36adc51ecfde)",
     "'As I have said, there is more in my life than PMTNM... Will have more time for this tomorrow morning.'",
     "S-03, S-04", "CRITICAL",
     "Sharon's complete substantive reply to a formal written request invoking Handbook Art. VIII §5 and three NM statutes. She addressed none of the five items, mentioned visiting conference venues and having students, and said she would have more time the next morning. She never followed up. Three weeks later she committed before attorney Kathy Black to place the items on the agenda. Three weeks after that she adjourned the board meeting without addressing a single one."),
    ("Jeanne Grealish — documented in Chip's Oct 14, 2025 'Silence, officer training, and concerns' letter to Sharon",
     "'[Chip's facility research efforts are] useless.'",
     "S-04, G-01", "HIGH",
     "Second documented dismissive characterization of VP work product alongside 'boring and irrelevant' for the Foundation newsletter. Both reported in writing to Sharon Kunitz. Sharon did not intervene in either case. Source: Gmail thread 199e369bfb1d7543."),
    ("Jeanne Grealish — documented in Chip's Oct 14, 2025 formal agenda letter to Sharon, Item 3",
     "'Why not have Laura choose someone she'd like to work with even if that person has not served on the Board?'",
     "G-01", "HIGH",
     "Grealish suggested the incoming President handpick the next VP outside the Handbook's formal nominating committee process. The nominating committee is led by the Immediate Past President (Larry Blind) — the same person who administered the biased TOTY process. Having Grealish suggest an alternative succession route places VP selection under informal control. Documented verbatim in VP's formal five-item agenda request, Gmail thread 199e36adc51ecfde."),
    ("Lawrence Blind — Oct 28, 2023 email to PMTNM teachers",
     "[Unilateral invitation to NMSM students to participate in PMTNM conference benefit concert]",
     "G-03, L-02", "MEDIUM",
     "No board approval, no recorded vote, no disclosure of conflict. President used organizational platform to promote his own school. Handbook silent on conference programming authority, creating governance gap. Pattern consistent with leadership's 'no legal mandates' operational style."),
    ("Lawrence Blind — Oct 12, 2023 text message to Chip Miller (NMSM context)",
     "'Chip, please do not contact any of your former students from our school by phone or email. This is not appropriate.'",
     "E-04, L-04", "HIGH",
     "Used NMSM authority to block teacher from complying with MTNA Code of Ethics requiring teacher participation in student transitions. When escalated to MTNA (Jul 2024), the organization could not enforce ethics standards because Larry designed NMSM handbook to conflict with MTNA code. Pattern: systematically uses organizational structure to block accountability mechanisms."),

]

for i, (speaker, quote, viol, severity, context) in enumerate(key_quotes, quote_start + 1):
    row_fill = severity_fills.get(severity, SLATE_FILL)
    vals = [speaker, quote, viol, severity, context]
    for j, val in enumerate(vals, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.fill = row_fill if j in [1, 4] else (PatternFill("solid", start_color="F9F9F9", fgColor="F9F9F9") if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", fgColor="FFFFFF"))
        c.font = cell_font(bold=(j in [1, 2]), italic=(j == 2))
        c.alignment = wrap()
        c.border = border
    ws4.row_dimensions[i].height = 72

# ── Save ──────────────────────────────────────────────────────────────────────
out = 'PMTNM_Violations_Database.xlsx'
wb.save(out)
print("Done:", out)
